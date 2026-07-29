from __future__ import annotations

import base64
import copy
import json
import io
import os
import shutil
import subprocess
import tempfile
import zipfile
import sys
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable
from xml.sax.saxutils import escape

from lxml import etree
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from reportlab.lib.utils import ImageReader

from delta_fakture_core import (
    APP_DIR,
    ASSETS_DIR,
    LOGO_FILE,
    TEMPLATE_XLSX,
    calculate_invoice_totals,
    format_currency,
    format_date,
    money_round,
    number_to_words_bg,
    safe_filename,
)


NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
XML_SPACE = "{http://www.w3.org/XML/1998/namespace}space"


def _powershell_executable() -> str:
    """Use an absolute Windows PowerShell path in packaged desktop builds."""
    windows_root = Path(os.environ.get("SystemRoot") or os.environ.get("WINDIR") or r"C:\Windows")
    candidate = windows_root / "System32" / "WindowsPowerShell" / "v1.0" / "powershell.exe"
    return str(candidate) if candidate.is_file() else "powershell.exe"


def _powershell_literal(value: Path | str) -> str:
    """Return one safe PowerShell string literal without JSON path escaping.

    ``json.dumps`` is correct for JavaScript but not for a PowerShell source
    string: a Windows path such as ``C:\\Temp`` becomes ``C:\\\\Temp`` and
    Excel COM can refuse to open it.  PowerShell escapes a single quote by
    doubling it, while backslashes must remain untouched.
    """
    return "'" + str(value).replace("'", "''") + "'"


PAGE_W, PAGE_H = A4
MARGIN_X = 14 * mm
MARGIN_TOP = 12 * mm
MARGIN_BOTTOM = 10 * mm
CONTENT_W = PAGE_W - 2 * MARGIN_X

GREEN_DARK = colors.HexColor("#245B41")
GREEN_MID = colors.HexColor("#2F7A56")
GREEN_LIGHT = colors.HexColor("#EAF4EC")
GREEN_LIGHTER = colors.HexColor("#F6FBF7")
GREEN_LINE = colors.HexColor("#B9D5C2")
TEXT_DARK = colors.HexColor("#2C3E37")
TEXT_MUTED = colors.HexColor("#6A7A72")

MAX_VISIBLE_ITEMS = 18
INVOICE_PRINT_AREA = "$A$1:$K$62"

# The supplied workbook is Bulgarian by default.  These labels make the issued
# Excel/PDF document independent from the language used in the application.
# Free-text line descriptions remain exactly as entered by the user: OpsNest
# must not silently machine-translate contractual wording on a legal document.
INVOICE_DOCUMENT_TEXT: dict[str, dict[str, str]] = {
    "bg": {
        "tax_document": "ДАНЪЧЕН ДОКУМЕНТ", "invoice": "ФАКТУРА", "construction": "СТРОИТЕЛНО-МОНТАЖНИ РАБОТИ", "currency": "ВАЛУТА",
        "issue_date": "ДАТА НА ИЗДАВАНЕ", "tax_event": "ДАТА НА ДАНЪЧНОТО СЪБИТИЕ", "due_date": "ПАДЕЖ", "payment_method": "НАЧИН НА ПЛАЩАНЕ",
        "project": "ОБЕКТ / ПРОЕКТ", "site": "АДРЕС НА ОБЕКТА", "contract": "ДОГОВОР №", "protocol": "ПРОТОКОЛ / АКТ ОБР. 19 №",
        "period_from": "ПЕРИОД НА СМР — ОТ", "period_to": "ДО", "reference": "ПОРЪЧКА / РЕФЕРЕНЦИЯ", "issue_place": "МЯСТО НА ИЗДАВАНЕ",
        "supplier": "ИЗПЪЛНИТЕЛ / ДОСТАВЧИК", "customer": "ВЪЗЛОЖИТЕЛ / ПОЛУЧАТЕЛ", "company": "Фирма", "eik": "ЕИК / Булстат", "vat_no": "ИН по ДДС", "address": "Адрес", "manager": "МОЛ / Управител", "contact": "Телефон / Имейл",
        "breakdown": "РАЗБИВКА НА ИЗВЪРШЕНИТЕ СМР, МАТЕРИАЛИ И РАЗХОДИ", "category": "Категория", "description": "Вид СМР / описание", "unit": "Мярка", "quantity": "Количество", "unit_price": "Ед. цена без ДДС", "discount": "Отстъпка %", "net": "Стойност без ДДС", "vat": "ДДС {vat}%", "gross": "Общо с ДДС", "code": "Код / етап",
        "notes": "ОСНОВАНИЕ И БЕЛЕЖКИ", "subtotal": "СТОЙНОСТ БЕЗ ДДС", "adjustment": "ДОПЪЛНИТЕЛНА ОТСТЪПКА / КОРЕКЦИЯ БЕЗ ДДС", "tax_base": "ДАНЪЧНА ОСНОВА", "total": "ОБЩО С ДДС", "retention": "ГАРАНЦИОННО УДЪРЖАНЕ ПО ДОГОВОР", "advance": "ПРИСПАДНАТ АВАНС (С ДДС)", "paid": "ПЛАТЕНО", "balance": "ОСТАТЪК ЗА ПЛАЩАНЕ", "words": "СУМА С ДУМИ", "reference_bgn": "СПРАВОЧНО В BGN", "payment_deadline": "СРОК ЗА ПЛАЩАНЕ", "bank_details": "БАНКОВИ ДАННИ ЗА ПЛАЩАНЕ", "bank": "БАНКА", "payment_reference": "Основание за плащане: фактура № {number}", "bank_hint": "Моля, посочете номера на фактурата при банков превод.", "prepared": "Съставил:", "received": "Получил / Одобрил:",
        "default_note": "Извършени строително-монтажни работи съгласно договор, количествено-стойностна сметка и/или подписан протокол.",
    },
    "sr": {
        "tax_document": "PORESKI DOKUMENT", "invoice": "FAKTURA", "construction": "GRAĐEVINSKI RADOVI", "currency": "VALUTA",
        "issue_date": "DATUM IZDAVANJA", "tax_event": "DATUM PORESKOG DOGAĐAJA", "due_date": "ROK PLAĆANJA", "payment_method": "NAČIN PLAĆANJA",
        "project": "OBJEKAT / PROJEKAT", "site": "ADRESA OBJEKTA", "contract": "UGOVOR BR.", "protocol": "PROTOKOL / AKT 19 BR.",
        "period_from": "PERIOD RADOVA — OD", "period_to": "DO", "reference": "NARUDŽBENICA / REFERENCA", "issue_place": "MESTO IZDAVANJA",
        "supplier": "IZVOĐAČ / DOBAVLJAČ", "customer": "NARUČILAC / PRIMALAC", "company": "Firma", "eik": "Matični broj", "vat_no": "PDV broj", "address": "Adresa", "manager": "Odgovorno lice", "contact": "Telefon / e-mail",
        "breakdown": "PREGLED RADOVA, MATERIJALA I TROŠKOVA", "category": "Kategorija", "description": "Opis", "unit": "Jedinica", "quantity": "Količina", "unit_price": "Jed. cena bez PDV-a", "discount": "Popust %", "net": "Vrednost bez PDV-a", "vat": "PDV {vat}%", "gross": "Ukupno sa PDV-om", "code": "Šifra / etapa",
        "notes": "OSNOV I NAPOMENE", "subtotal": "VREDNOST BEZ PDV-a", "adjustment": "DODATNI POPUST / KOREKCIJA BEZ PDV-a", "tax_base": "PORESKA OSNOVICA", "total": "UKUPNO SA PDV-om", "retention": "GARANCIJSKO ZADRŽAVANJE PO UGOVORU", "advance": "ODBIJENI AVANS (SA PDV-om)", "paid": "PLAĆENO", "balance": "OSTATAK ZA PLAĆANJE", "words": "IZNOS SLOVIMA", "reference_bgn": "INFORMATIVNO U BGN", "payment_deadline": "ROK PLAĆANJA", "bank_details": "BANKOVNI PODACI ZA PLAĆANJE", "bank": "BANKA", "payment_reference": "Poziv na broj: faktura br. {number}", "bank_hint": "Molimo navedite broj fakture pri plaćanju.", "prepared": "Sastavio:", "received": "Primio / odobrio:",
        "default_note": "Izvedeni građevinsko-montažni radovi prema ugovoru, obračunu količina i/ili potpisanom protokolu.",
    },
    "en": {
        "tax_document": "TAX DOCUMENT", "invoice": "INVOICE", "construction": "CONSTRUCTION WORKS", "currency": "CURRENCY",
        "issue_date": "ISSUE DATE", "tax_event": "TAX POINT DATE", "due_date": "DUE DATE", "payment_method": "PAYMENT METHOD",
        "project": "SITE / PROJECT", "site": "SITE ADDRESS", "contract": "CONTRACT NO.", "protocol": "PROTOCOL / ACT 19 NO.",
        "period_from": "WORK PERIOD — FROM", "period_to": "TO", "reference": "ORDER / REFERENCE", "issue_place": "PLACE OF ISSUE",
        "supplier": "SUPPLIER", "customer": "CUSTOMER", "company": "Company", "eik": "Company ID", "vat_no": "VAT number", "address": "Address", "manager": "Responsible person", "contact": "Phone / e-mail",
        "breakdown": "BREAKDOWN OF WORKS, MATERIALS AND COSTS", "category": "Category", "description": "Description", "unit": "Unit", "quantity": "Quantity", "unit_price": "Unit price excl. VAT", "discount": "Discount %", "net": "Net amount", "vat": "VAT {vat}%", "gross": "Total incl. VAT", "code": "Code / stage",
        "notes": "BASIS AND NOTES", "subtotal": "VALUE EXCL. VAT", "adjustment": "ADDITIONAL DISCOUNT / ADJUSTMENT EXCL. VAT", "tax_base": "TAX BASE", "total": "TOTAL INCL. VAT", "retention": "CONTRACT RETENTION", "advance": "DEDUCTED ADVANCE (INCL. VAT)", "paid": "PAID", "balance": "BALANCE DUE", "words": "AMOUNT IN WORDS", "reference_bgn": "REFERENCE IN BGN", "payment_deadline": "PAYMENT DEADLINE", "bank_details": "BANK DETAILS FOR PAYMENT", "bank": "BANK", "payment_reference": "Payment reference: invoice no. {number}", "bank_hint": "Please quote the invoice number when making payment.", "prepared": "Prepared by:", "received": "Received / approved by:",
        "default_note": "Construction and installation works performed in accordance with the contract, bill of quantities and/or signed protocol.",
    },
}


def invoice_document_language(invoice: dict[str, Any]) -> str:
    language = str(invoice.get("document_language") or "").strip().lower()
    if language in INVOICE_DOCUMENT_TEXT:
        return language
    company = invoice.get("company") or {}
    return "bg" if str(company.get("country_code") or "").upper() == "BG" else "sr"


def invoice_document_text(invoice: dict[str, Any]) -> dict[str, str]:
    return INVOICE_DOCUMENT_TEXT[invoice_document_language(invoice)]


def _translated_category(value: Any, language: str) -> str:
    source = str(value or "").strip()
    normalized = source.casefold().strip(" .")
    canonical = {
        "rad": "labour", "labor": "labour", "labour": "labour", "труд": "labour",
        "materijal": "materials", "materijali": "materials", "materials": "materials", "материали": "materials",
        "mehanizacija": "machinery", "machinery": "machinery", "механизация": "machinery",
        "transport": "transport", "транспорт": "transport",
        "ostalo": "other", "drugi": "other", "other": "other", "други": "other",
        "ugovorni avans": "advance", "avans": "advance", "advance": "advance", "аванc": "advance", "аванс": "advance",
    }.get(normalized)
    translated = {
        "labour": {"sr": "Rad", "bg": "Труд", "en": "Labour"},
        "materials": {"sr": "Materijal", "bg": "Материали", "en": "Materials"},
        "machinery": {"sr": "Mehanizacija", "bg": "Механизация", "en": "Machinery"},
        "transport": {"sr": "Transport", "bg": "Транспорт", "en": "Transport"},
        "other": {"sr": "Ostalo", "bg": "Други", "en": "Other"},
        "advance": {"sr": "Ugovorni avans", "bg": "Договорен аванс", "en": "Contract advance"},
    }
    return translated.get(canonical or "", {}).get(language, source)


def _translated_unit(value: Any, language: str) -> str:
    """Translate only standardized units, never a free-form unit supplied by a user."""
    source = str(value or "").strip()
    canonical = {
        "kom": "piece", "kom.": "piece", "pcs": "piece", "pc": "piece", "бр": "piece", "бр.": "piece",
        "usl": "service", "usl.": "service", "usluga": "service", "service": "service", "усл": "service", "усл.": "service",
        "sat": "hour", "h": "hour", "hour": "hour", "час": "hour", "ч": "hour",
        "dan": "day", "day": "day", "ден": "day",
    }.get(source.casefold())
    translated = {
        "piece": {"sr": "kom.", "bg": "бр.", "en": "pcs"},
        "service": {"sr": "usl.", "bg": "усл.", "en": "service"},
        "hour": {"sr": "sat", "bg": "час", "en": "hour"},
        "day": {"sr": "dan", "bg": "ден", "en": "day"},
    }
    return translated.get(canonical or "", {}).get(language, source)


def _translated_payment_method(value: Any, language: str) -> str:
    """Keep a selected payment method readable in the document language."""
    source = str(value or "").strip()
    canonical = {
        "banka": "bank", "bankarski transfer": "bank", "bank transfer": "bank", "банков превод": "bank",
        "gotovina": "cash", "kasa": "cash", "cash": "cash", "в брой": "cash",
        "kartica": "card", "card": "card", "карта": "card",
        "kompenzacija": "offset", "offset": "offset", "прихващане": "offset",
    }.get(source.casefold())
    translated = {
        "bank": {"sr": "Banka", "bg": "Банков превод", "en": "Bank transfer"},
        "cash": {"sr": "Gotovina", "bg": "В брой", "en": "Cash"},
        "card": {"sr": "Kartica", "bg": "Карта", "en": "Card"},
        "offset": {"sr": "Kompenzacija", "bg": "Прихващане", "en": "Set-off"},
    }
    return translated.get(canonical or "", {}).get(language, source)


def _translated_standard_description(value: Any, language: str, invoice_kind: Any = "") -> str:
    """Translate OpsNest's own financial terms while preserving custom legal wording.

    For example, the automatically created ``Avans 20%`` is a system term,
    unlike an item description the accountant typed manually.
    """
    source = str(value or "").strip()
    normalized = source.casefold()
    advance_prefixes = ("avans", "ugovorni avans", "advance", "договорен аванс", "аванс")
    for prefix in advance_prefixes:
        if normalized == prefix or normalized.startswith(prefix + " "):
            suffix = source[len(prefix):].strip()
            label = {"sr": "Avans", "bg": "Аванс", "en": "Advance payment"}.get(language, "Avans")
            suffix_prefixes = {
                "sr": (("po ugovoru", "po ugovoru"),),
                "bg": (("po ugovoru", "по договор"), ("prema ugovoru", "съгласно договор")),
                "en": (("po ugovoru", "under contract"), ("prema ugovoru", "under contract")),
            }
            for original, localized in suffix_prefixes.get(language, ()):
                position = suffix.casefold().find(original)
                if position >= 0:
                    suffix = suffix[:position] + localized + suffix[position + len(original):]
                    break
            return f"{label}{(' ' + suffix) if suffix else ''}"
    if str(invoice_kind or "").strip().lower() == "advance" and not source:
        return {"sr": "Avans", "bg": "Аванс", "en": "Advance payment"}.get(language, "Avans")
    return source


def _credit_note_document_language(note: dict[str, Any]) -> str:
    source = _credit_note_source(note)
    language = str(source.get("document_language") or note.get("document_language") or "").strip().lower()
    if language in INVOICE_DOCUMENT_TEXT:
        return language
    company = note.get("company") if isinstance(note.get("company"), dict) else {}
    return "bg" if str(company.get("country_code") or "").upper() == "BG" else "sr"


def _credit_note_text(note: dict[str, Any]) -> dict[str, str]:
    language = _credit_note_document_language(note)
    labels = {
        "sr": {
            "title": "KREDITNO ODOBRENJE", "subtitle": "Formalni dokument uz izdatu fakturu", "number": "BROJ ODOBRENJA", "issue_date": "DATUM IZDAVANJA", "source_invoice": "IZVORNA FAKTURA", "source_date": "DATUM FAKTURE", "project": "PROJEKAT", "currency": "VALUTA", "supplier": "IZDAVALAC", "customer": "PRIMALAC", "reason": "OSNOV KOREKCIJE", "net": "Osnovica bez PDV-a", "vat": "PDV", "total": "Ukupno odobrenje", "linked_invoice": "Povezana faktura", "archive_note": "Dokument je vezan za izvornu fakturu i evidentirani povraćaj. Sačuvati zajedno sa poreskom evidencijom projekta.",
        },
        "bg": {
            "title": "КРЕДИТНО ИЗВЕСТИЕ", "subtitle": "Официален документ към издадена фактура", "number": "НОМЕР НА КРЕДИТНОТО ИЗВЕСТИЕ", "issue_date": "ДАТА НА ИЗДАВАНЕ", "source_invoice": "ОРИГИНАЛНА ФАКТУРА", "source_date": "ДАТА НА ФАКТУРАТА", "project": "ОБЕКТ / ПРОЕКТ", "currency": "ВАЛУТА", "supplier": "ДОСТАВЧИК", "customer": "ПОЛУЧАТЕЛ", "reason": "ОСНОВАНИЕ ЗА КОРЕКЦИЯ", "net": "Данъчна основа", "vat": "ДДС", "total": "Общо кредитно известие", "linked_invoice": "Свързана фактура", "archive_note": "Документът е свързан с оригиналната фактура и регистрираното възстановяване. Съхранявайте го заедно с данъчната документация на проекта.",
        },
        "en": {
            "title": "CREDIT NOTE", "subtitle": "Formal document linked to an issued invoice", "number": "CREDIT NOTE NUMBER", "issue_date": "ISSUE DATE", "source_invoice": "ORIGINAL INVOICE", "source_date": "INVOICE DATE", "project": "PROJECT", "currency": "CURRENCY", "supplier": "SUPPLIER", "customer": "CUSTOMER", "reason": "REASON FOR CORRECTION", "net": "Net amount", "vat": "VAT", "total": "Credit note total", "linked_invoice": "Linked invoice", "archive_note": "This document is linked to the original invoice and the recorded refund. Keep it with the project's tax records.",
        },
    }
    return labels[language]

# Reports follow the application language by default, but the caller can set
# report_language on a report payload to prepare a package for another reader.
REPORT_LANGUAGE_CODES = {"sr", "en", "de", "bg", "ru"}
REPORT_TEXT: dict[str, dict[str, str]] = {
    "sr": {
        "vat_file": "PDV_evidencija", "accountant_file": "izvoz_za_knjigovodju",
        "vat_title": "PDV EVIDENCIJA PROJEKTA", "vat_subtitle": "Radni izvoz za knjigovođu - nije XML fajl niti direktna prijava za NRA",
        "accountant_title": "IZVOZ PROJEKTA ZA KNJIGOVOĐU", "accountant_note": "Radni paket za knjigovođu: izlazne i ulazne račune, uplate/povraćaje, PDV, odobrenja i storna proverite pre knjiženja. Ovo nije direktan NRA XML izvoz.",
        "company": "Firma", "project": "Projekat", "site": "Gradilište", "period": "Period", "generated": "Generisano",
        "date": "Datum", "type": "Tip", "document_type": "Tip dokumenta", "document_number": "Broj dokumenta", "invoice": "Faktura", "customer": "Kupac", "partner": "Partner", "partner_vat": "PDV broj partnera", "description": "Opis", "payment_method_note": "Način / napomena",
        "net_base": "Osnovica bez PDV-a", "net_amount": "Osnovica / iznos", "vat": "PDV", "total": "Ukupno", "summary": "Pregled",
        "output_vat": "Izlazni PDV", "input_vat": "Ulazni PDV", "vat_payable": "PDV za uplatu / pretplatu", "output_ledger": "Izlazni PDV", "input_ledger": "Ulazni PDV",
        "outgoing_invoices": "Izlazne fakture", "incoming_bills": "Ulazni računi / troškovi", "payments": "Uplate", "refunds": "Povraćaji", "collected": "Naplaćeno", "payments_refunds": "Uplate i povraćaji", "credit_cancellations": "Odobrenja i storna",
        "control": "Kontrola", "control_title": "KONTROLA PRE IZVOZA ZA KNJIGOVOĐU", "foreign_currency": "Stavke van EUR", "missing_date": "Stavke bez datuma", "currency": "Valuta", "number": "Broj", "none": "Nema", "empty_period": "Nema stavki u izabranom periodu.",
        "issued_invoice": "Izdana faktura", "incoming_bill": "Ulazni račun", "outgoing_bill": "Izlazni račun", "credit_note": "Kreditno odobrenje", "cancelled_invoice": "Stornirana faktura", "payment": "Uplata", "refund": "Povraćaj uplate",
        "vat_working_note": "Ovo je radna evidencija za proveru sa knjigovođom. Nije XML fajl niti direktna prijava za NRA. Izdane fakture, kreditna odobrenja i ulazni računi su detaljno prikazani na zasebnim listovima.",
        "vat_control_warning": "Kontrola: {foreign} stavki van EUR i {missing} stavki bez datuma nisu uključene u zbir PDV-a. Proverite list '{control}' u Excel kopiji pre knjiženja.",
        "accountant_control_warning": "Kontrola: {foreign} stavki van EUR i {missing} stavki bez datuma nisu uključene u PDV zbir. Detalji su na listu '{control}' u Excel kopiji.",
        "vat_footer": "OpsNest - PDV evidencija projekta za internu proveru i knjigovođu", "accountant_footer": "OpsNest - kompletan izvoz projekta za knjigovođu", "page": "Strana",
        "sheet_summary": "Sažetak", "sheet_output": "Izlazni PDV", "sheet_input": "Ulazni PDV", "sheet_control": "Kontrola", "sheet_outgoing": "Izlazne fakture", "sheet_incoming": "Ulazni računi", "sheet_payments": "Uplate i povraćaji", "sheet_corrections": "Odobrenja i storna", "sheet_vat": "PDV pregled",
    },
    "en": {
        "vat_file": "VAT_ledger", "accountant_file": "accountant_export", "vat_title": "PROJECT VAT LEDGER", "vat_subtitle": "Working export for the accountant - not an XML file or a direct NRA filing", "accountant_title": "PROJECT EXPORT FOR ACCOUNTING", "accountant_note": "Working package for the accountant: review outgoing and incoming bills, payments/refunds, VAT, credit notes and cancellations before posting. This is not a direct NRA XML export.",
        "company": "Company", "project": "Project", "site": "Site", "period": "Period", "generated": "Generated", "date": "Date", "type": "Type", "document_type": "Document type", "document_number": "Document number", "invoice": "Invoice", "customer": "Customer", "partner": "Partner", "partner_vat": "Partner VAT number", "description": "Description", "payment_method_note": "Method / note", "net_base": "Net amount", "net_amount": "Net amount", "vat": "VAT", "total": "Total", "summary": "Overview", "output_vat": "Output VAT", "input_vat": "Input VAT", "vat_payable": "VAT payable / refundable", "output_ledger": "Output VAT", "input_ledger": "Input VAT", "outgoing_invoices": "Outgoing invoices", "incoming_bills": "Incoming bills / costs", "payments": "Payments", "refunds": "Refunds", "collected": "Collected", "payments_refunds": "Payments and refunds", "credit_cancellations": "Credit notes and cancellations", "control": "Control", "control_title": "PRE-EXPORT CONTROL FOR ACCOUNTING", "foreign_currency": "Non-EUR items", "missing_date": "Items without date", "currency": "Currency", "number": "Number", "none": "None", "empty_period": "No items in the selected period.", "issued_invoice": "Issued invoice", "incoming_bill": "Incoming bill", "outgoing_bill": "Outgoing bill", "credit_note": "Credit note", "cancelled_invoice": "Cancelled invoice", "payment": "Payment", "refund": "Payment refund", "vat_working_note": "This is a working ledger for review with the accountant. It is not an XML file or a direct NRA filing. Issued invoices, credit notes and incoming bills are detailed on separate sheets.", "vat_control_warning": "Control: {foreign} non-EUR items and {missing} items without a date are excluded from VAT totals. Review the '{control}' sheet in the Excel copy before posting.", "accountant_control_warning": "Control: {foreign} non-EUR items and {missing} items without a date are excluded from VAT totals. Details are on the '{control}' sheet in the Excel copy.", "vat_footer": "OpsNest - project VAT ledger for internal review and accounting", "accountant_footer": "OpsNest - complete project export for accounting", "page": "Page", "sheet_summary": "Summary", "sheet_output": "Output VAT", "sheet_input": "Input VAT", "sheet_control": "Control", "sheet_outgoing": "Outgoing invoices", "sheet_incoming": "Incoming bills", "sheet_payments": "Payments and refunds", "sheet_corrections": "Credit notes and cancellations", "sheet_vat": "VAT overview",
    },
    "de": {
        "vat_file": "USt_Uebersicht", "accountant_file": "Buchhaltungsexport", "vat_title": "UST-ÜBERSICHT DES PROJEKTS", "vat_subtitle": "Arbeitsauszug für die Buchhaltung - keine XML-Datei und keine direkte NRA-Meldung", "accountant_title": "PROJEKTEXPORT FÜR DIE BUCHHALTUNG", "accountant_note": "Arbeitspaket für die Buchhaltung: Ausgangs- und Eingangsrechnungen, Zahlungen/Erstattungen, USt., Gutschriften und Stornierungen vor der Buchung prüfen. Dies ist kein direkter NRA-XML-Export.",
        "company": "Firma", "project": "Projekt", "site": "Baustelle", "period": "Zeitraum", "generated": "Erstellt", "date": "Datum", "type": "Typ", "document_type": "Dokumenttyp", "document_number": "Dokumentnummer", "invoice": "Rechnung", "customer": "Kunde", "partner": "Partner", "partner_vat": "USt-IdNr. Partner", "description": "Beschreibung", "payment_method_note": "Zahlungsart / Hinweis", "net_base": "Nettobetrag", "net_amount": "Nettobetrag", "vat": "USt.", "total": "Gesamt", "summary": "Übersicht", "output_vat": "Ausgangs-USt.", "input_vat": "Vorsteuer", "vat_payable": "USt.-Zahllast / Erstattung", "output_ledger": "Ausgangs-USt.", "input_ledger": "Vorsteuer", "outgoing_invoices": "Ausgangsrechnungen", "incoming_bills": "Eingangsrechnungen / Kosten", "payments": "Zahlungen", "refunds": "Erstattungen", "collected": "Eingezogen", "payments_refunds": "Zahlungen und Erstattungen", "credit_cancellations": "Gutschriften und Stornierungen", "control": "Kontrolle", "control_title": "KONTROLLE VOR BUCHHALTUNGSEXPORT", "foreign_currency": "Posten außerhalb EUR", "missing_date": "Posten ohne Datum", "currency": "Währung", "number": "Nummer", "none": "Keine", "empty_period": "Keine Posten im ausgewählten Zeitraum.", "issued_invoice": "Ausgestellte Rechnung", "incoming_bill": "Eingangsrechnung", "outgoing_bill": "Ausgangsrechnung", "credit_note": "Gutschrift", "cancelled_invoice": "Stornierte Rechnung", "payment": "Zahlung", "refund": "Zahlungserstattung", "vat_working_note": "Dies ist eine Arbeitsübersicht zur Prüfung mit der Buchhaltung. Sie ist keine XML-Datei und keine direkte NRA-Meldung. Ausgestellte Rechnungen, Gutschriften und Eingangsrechnungen sind auf separaten Blättern aufgeführt.", "vat_control_warning": "Kontrolle: {foreign} Posten außerhalb EUR und {missing} Posten ohne Datum sind nicht in den USt.-Summen enthalten. Vor der Buchung das Blatt '{control}' in der Excel-Kopie prüfen.", "accountant_control_warning": "Kontrolle: {foreign} Posten außerhalb EUR und {missing} Posten ohne Datum sind nicht in den USt.-Summen enthalten. Details stehen auf dem Blatt '{control}' in der Excel-Kopie.", "vat_footer": "OpsNest - USt.-Übersicht des Projekts für interne Prüfung und Buchhaltung", "accountant_footer": "OpsNest - vollständiger Projektexport für die Buchhaltung", "page": "Seite", "sheet_summary": "Zusammenfassung", "sheet_output": "Ausgangs-USt.", "sheet_input": "Vorsteuer", "sheet_control": "Kontrolle", "sheet_outgoing": "Ausgangsrechnungen", "sheet_incoming": "Eingangsrechnungen", "sheet_payments": "Zahlungen und Erstattungen", "sheet_corrections": "Gutschriften und Stornierungen", "sheet_vat": "USt.-Übersicht",
    },
    "bg": {
        "vat_file": "DDS_spravka", "accountant_file": "iznos_za_schetovodstvo", "vat_title": "ДДС СПРАВКА НА ПРОЕКТА", "vat_subtitle": "Работен износ за счетоводителя - не е XML файл или директна декларация към NRA", "accountant_title": "ИЗНОС НА ПРОЕКТА ЗА СЧЕТОВОДСТВОТО", "accountant_note": "Работен пакет за счетоводителя: проверете изходящи и входящи фактури, плащания/възстановявания, ДДС, кредитни известия и сторна преди осчетоводяване. Това не е директен NRA XML износ.",
        "company": "Фирма", "project": "Проект", "site": "Обект", "period": "Период", "generated": "Генерирано", "date": "Дата", "type": "Тип", "document_type": "Тип документ", "document_number": "Номер на документ", "invoice": "Фактура", "customer": "Клиент", "partner": "Партньор", "partner_vat": "ДДС номер на партньора", "description": "Описание", "payment_method_note": "Начин / бележка", "net_base": "Данъчна основа", "net_amount": "Сума без ДДС", "vat": "ДДС", "total": "Общо", "summary": "Преглед", "output_vat": "Изходящ ДДС", "input_vat": "Входящ ДДС", "vat_payable": "ДДС за плащане / възстановяване", "output_ledger": "Изходящ ДДС", "input_ledger": "Входящ ДДС", "outgoing_invoices": "Изходящи фактури", "incoming_bills": "Входящи фактури / разходи", "payments": "Плащания", "refunds": "Възстановявания", "collected": "Получено", "payments_refunds": "Плащания и възстановявания", "credit_cancellations": "Кредитни известия и сторна", "control": "Контрол", "control_title": "КОНТРОЛ ПРЕДИ ИЗНОС ЗА СЧЕТОВОДСТВО", "foreign_currency": "Позиции извън EUR", "missing_date": "Позиции без дата", "currency": "Валута", "number": "Номер", "none": "Няма", "empty_period": "Няма позиции за избрания период.", "issued_invoice": "Издадена фактура", "incoming_bill": "Входяща фактура", "outgoing_bill": "Изходяща фактура", "credit_note": "Кредитно известие", "cancelled_invoice": "Сторнирана фактура", "payment": "Плащане", "refund": "Възстановяване на плащане", "vat_working_note": "Това е работна справка за проверка със счетоводителя. Не е XML файл или директна декларация към NRA. Издадените фактури, кредитните известия и входящите фактури са на отделни листове.", "vat_control_warning": "Контрол: {foreign} позиции извън EUR и {missing} позиции без дата не са включени в сумите за ДДС. Проверете листа '{control}' в Excel копието преди осчетоводяване.", "accountant_control_warning": "Контрол: {foreign} позиции извън EUR и {missing} позиции без дата не са включени в сумите за ДДС. Детайлите са в листа '{control}' в Excel копието.", "vat_footer": "OpsNest - ДДС справка на проекта за вътрешна проверка и счетоводство", "accountant_footer": "OpsNest - пълен износ на проекта за счетоводството", "page": "Страница", "sheet_summary": "Обобщение", "sheet_output": "Изходящ ДДС", "sheet_input": "Входящ ДДС", "sheet_control": "Контрол", "sheet_outgoing": "Изходящи фактури", "sheet_incoming": "Входящи фактури", "sheet_payments": "Плащания и възстановявания", "sheet_corrections": "Кредитни известия и сторна", "sheet_vat": "ДДС преглед",
    },
    "ru": {
        "vat_file": "uchet_NDS", "accountant_file": "eksport_dlya_buhgalterii", "vat_title": "ЖУРНАЛ НДС ПО ПРОЕКТУ", "vat_subtitle": "Рабочий экспорт для бухгалтера - не XML-файл и не прямая подача в NRA", "accountant_title": "ЭКСПОРТ ПРОЕКТА ДЛЯ БУХГАЛТЕРИИ", "accountant_note": "Рабочий пакет для бухгалтера: проверьте исходящие и входящие счета, платежи/возвраты, НДС, кредитовые ноты и сторно перед проводкой. Это не прямой NRA XML-экспорт.",
        "company": "Компания", "project": "Проект", "site": "Строительная площадка", "period": "Период", "generated": "Создано", "date": "Дата", "type": "Тип", "document_type": "Тип документа", "document_number": "Номер документа", "invoice": "Счет", "customer": "Клиент", "partner": "Контрагент", "partner_vat": "Номер НДС контрагента", "description": "Описание", "payment_method_note": "Способ / примечание", "net_base": "Сумма без НДС", "net_amount": "Сумма без НДС", "vat": "НДС", "total": "Итого", "summary": "Обзор", "output_vat": "Исходящий НДС", "input_vat": "Входящий НДС", "vat_payable": "НДС к уплате / возврату", "output_ledger": "Исходящий НДС", "input_ledger": "Входящий НДС", "outgoing_invoices": "Исходящие счета", "incoming_bills": "Входящие счета / расходы", "payments": "Платежи", "refunds": "Возвраты", "collected": "Получено", "payments_refunds": "Платежи и возвраты", "credit_cancellations": "Кредитовые ноты и сторно", "control": "Контроль", "control_title": "КОНТРОЛЬ ПЕРЕД ЭКСПОРТОМ ДЛЯ БУХГАЛТЕРИИ", "foreign_currency": "Позиции вне EUR", "missing_date": "Позиции без даты", "currency": "Валюта", "number": "Номер", "none": "Нет", "empty_period": "Нет позиций за выбранный период.", "issued_invoice": "Выставленный счет", "incoming_bill": "Входящий счет", "outgoing_bill": "Исходящий счет", "credit_note": "Кредитовая нота", "cancelled_invoice": "Сторнированный счет", "payment": "Платеж", "refund": "Возврат платежа", "vat_working_note": "Это рабочий журнал для проверки с бухгалтером. Он не является XML-файлом или прямой подачей в NRA. Выставленные счета, кредитовые ноты и входящие счета приведены на отдельных листах.", "vat_control_warning": "Контроль: {foreign} позиций вне EUR и {missing} позиций без даты исключены из сумм НДС. Проверьте лист '{control}' в копии Excel перед проводкой.", "accountant_control_warning": "Контроль: {foreign} позиций вне EUR и {missing} позиций без даты исключены из сумм НДС. Подробности находятся на листе '{control}' в копии Excel.", "vat_footer": "OpsNest - журнал НДС проекта для внутренней проверки и бухгалтерии", "accountant_footer": "OpsNest - полный экспорт проекта для бухгалтерии", "page": "Страница", "sheet_summary": "Сводка", "sheet_output": "Исходящий НДС", "sheet_input": "Входящий НДС", "sheet_control": "Контроль", "sheet_outgoing": "Исходящие счета", "sheet_incoming": "Входящие счета", "sheet_payments": "Платежи и возвраты", "sheet_corrections": "Кредитовые ноты и сторно", "sheet_vat": "Обзор НДС",
    },
}

# EIK / BULSTAT is a company identifier, so keep its name aligned with the
# language of the exported report as well as the rest of the heading fields.
REPORT_COMPANY_IDENTIFIER_LABELS = {
    "sr": "EIK / BULSTAT",
    "en": "Company ID (EIK / BULSTAT)",
    "de": "Unternehmens-ID (EIK / BULSTAT)",
    "bg": "ЕИК / БУЛСТАТ",
    "ru": "ЕИК / БУЛСТАТ",
}


def normalize_report_language(value: Any) -> str:
    code = str(value or "sr").strip().lower()
    return code if code in REPORT_LANGUAGE_CODES else "sr"


def company_identifier_label(language: Any = "sr") -> str:
    return REPORT_COMPANY_IDENTIFIER_LABELS[normalize_report_language(language)]


def _report_language(report: dict[str, Any]) -> str:
    return normalize_report_language(report.get("report_language"))


def report_text(key: str, language: Any = "sr") -> str:
    code = normalize_report_language(language)
    return REPORT_TEXT.get(code, REPORT_TEXT["sr"]).get(key, REPORT_TEXT["sr"].get(key, key))


def _localized_document_type(value: Any, language: Any) -> str:
    source = str(value or "").strip()
    if source.startswith("Povraćaj"):
        return report_text("refund", language)
    keys = {
        "Izdana faktura": "issued_invoice",
        "Ulazni račun": "incoming_bill",
        "Izlazni račun": "outgoing_bill",
        "Kreditno odobrenje": "credit_note",
        "Stornirana faktura": "cancelled_invoice",
        "Uplata": "payment",
        "Povraćaj uplate": "refund",
    }
    return report_text(keys[source], language) if source in keys else source

LOCAL_RUNTIME_NODE_EXE = APP_DIR / "runtime" / "node" / "bin" / "node.exe"
LOCAL_ARTIFACT_TOOL = APP_DIR / "runtime" / "node" / "node_modules" / "@oai" / "artifact-tool" / "dist" / "artifact_tool.mjs"
CODEx_RUNTIME_NODE_EXE = Path(r"C:\Users\49162\.cache\codex-runtimes\codex-primary-runtime\dependencies\node\bin\node.exe")
CODEx_ARTIFACT_TOOL = Path(
    r"C:\Users\49162\.cache\codex-runtimes\codex-primary-runtime\dependencies\node\node_modules\@oai\artifact-tool\dist\artifact_tool.mjs"
)


def _register_fonts() -> tuple[str, str]:
    candidates = [
        (Path("C:/Windows/Fonts/arial.ttf"), Path("C:/Windows/Fonts/arialbd.ttf")),
        (Path("C:/Windows/Fonts/Calibri.ttf"), Path("C:/Windows/Fonts/Calibrib.ttf")),
    ]
    for regular, bold in candidates:
        if regular.exists() and bold.exists():
            if "DeltaRegular" not in pdfmetrics.getRegisteredFontNames():
                pdfmetrics.registerFont(TTFont("DeltaRegular", str(regular)))
            if "DeltaBold" not in pdfmetrics.getRegisteredFontNames():
                pdfmetrics.registerFont(TTFont("DeltaBold", str(bold)))
            return "DeltaRegular", "DeltaBold"
    return "Helvetica", "Helvetica-Bold"


FONT_REGULAR, FONT_BOLD = _register_fonts()


def _short_text(value: Any, limit: int = 90) -> str:
    text = "" if value is None else str(value)
    text = " ".join(text.split())
    if len(text) <= limit:
        return text
    return text[: max(0, limit - 1)].rstrip() + "…"


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return format_date(value)
    return str(value)


def _number_value(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float, Decimal)):
        return float(value)
    try:
        return float(Decimal(str(value)))
    except Exception:
        return 0.0


def _find_cell(root: etree._Element, ref: str) -> etree._Element | None:
    cells = root.xpath(f".//main:c[@r='{ref}']", namespaces={"main": NS_MAIN})
    return cells[0] if cells else None


def _ensure_row(root: etree._Element, row_number: int) -> etree._Element:
    rows = root.xpath(f".//main:sheetData/main:row[@r='{row_number}']", namespaces={"main": NS_MAIN})
    if rows:
        return rows[0]
    sheet_data = root.xpath(".//main:sheetData", namespaces={"main": NS_MAIN})
    if not sheet_data:
        raise RuntimeError("worksheet missing sheetData")
    row = etree.Element(f"{{{NS_MAIN}}}row", r=str(row_number))
    sheet_data[0].append(row)
    return row


def _set_inline_text(root: etree._Element, ref: str, value: Any) -> None:
    cell = _find_cell(root, ref)
    if cell is None:
        row_num = int("".join(ch for ch in ref if ch.isdigit()))
        row = _ensure_row(root, row_num)
        cell = etree.SubElement(row, f"{{{NS_MAIN}}}c", r=ref)
    for child in list(cell):
        cell.remove(child)
    cell.attrib.pop("t", None)
    if value is None or value == "":
        return
    cell.attrib["t"] = "inlineStr"
    is_el = etree.SubElement(cell, f"{{{NS_MAIN}}}is")
    t_el = etree.SubElement(is_el, f"{{{NS_MAIN}}}t")
    text = str(value)
    if text.startswith(" ") or text.endswith(" ") or "  " in text or "\n" in text:
        t_el.set(XML_SPACE, "preserve")
    t_el.text = text


def _set_number(root: etree._Element, ref: str, value: Any) -> None:
    cell = _find_cell(root, ref)
    if cell is None:
        row_num = int("".join(ch for ch in ref if ch.isdigit()))
        row = _ensure_row(root, row_num)
        cell = etree.SubElement(row, f"{{{NS_MAIN}}}c", r=ref)
    for child in list(cell):
        cell.remove(child)
    if value is None or value == "":
        cell.attrib.pop("t", None)
        return
    cell.attrib["t"] = "n"
    v = etree.SubElement(cell, f"{{{NS_MAIN}}}v")
    if isinstance(value, bool):
        v.text = "1" if value else "0"
    else:
        v.text = f"{float(value):.15g}"


def _set_bool(root: etree._Element, ref: str, value: bool) -> None:
    cell = _find_cell(root, ref)
    if cell is None:
        row_num = int("".join(ch for ch in ref if ch.isdigit()))
        row = _ensure_row(root, row_num)
        cell = etree.SubElement(row, f"{{{NS_MAIN}}}c", r=ref)
    for child in list(cell):
        cell.remove(child)
    cell.attrib["t"] = "b"
    v = etree.SubElement(cell, f"{{{NS_MAIN}}}v")
    v.text = "1" if value else "0"


def _copy_zip_with_updates(template_path: Path, output_path: Path, updates: dict[str, dict[str, Any]]) -> None:
    with zipfile.ZipFile(template_path, "r") as zin:
        tmp_path = output_path.with_suffix(output_path.suffix + ".tmp")
        with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
            for info in zin.infolist():
                data = zin.read(info.filename)
                if info.filename in updates:
                    root = etree.fromstring(data)
                    for ref, value in updates[info.filename].items():
                        if isinstance(value, bool):
                            _set_bool(root, ref, value)
                        elif isinstance(value, (int, float, Decimal)):
                            _set_number(root, ref, value)
                        else:
                            _set_inline_text(root, ref, value)
                    data = etree.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=False)
                zout.writestr(info, data)
    if output_path.exists():
        output_path.unlink()
    tmp_path.replace(output_path)


def _company_display(company: dict[str, Any]) -> dict[str, str]:
    phone = _short_text(company.get("phone", ""))
    email = _short_text(company.get("email", ""))
    phone_email = " • ".join([part for part in [phone, email] if part]) or ""
    return {
        "name": _short_text(company.get("name", "")),
        "eik": _short_text(company.get("eik", "")),
        "vat": _short_text(company.get("vat_number", "")),
        "address": _short_text(company.get("address", "")),
        "director": _short_text(company.get("director_name", "")),
        "phone_email": phone_email,
        "bank_name": _short_text(company.get("bank_name", "")),
        "iban": _short_text(company.get("iban", "")),
        "bic": _short_text(company.get("bic", "")),
        "bank_line": _short_text(company.get("bank_name", "")),
    }


def build_invoice_xlsx_updates(invoice: dict[str, Any]) -> dict[str, dict[str, Any]]:
    company = invoice.get("company", {})
    display = _company_display(company)
    text = invoice_document_text(invoice)
    language = invoice_document_language(invoice)
    customer_phone = _short_text(invoice.get("customer_phone", ""))
    customer_email = _short_text(invoice.get("customer_email", ""))
    customer_phone_email = " • ".join([part for part in [customer_phone, customer_email] if part]) or ""

    totals = calculate_invoice_totals(
        invoice.get("items", []),
        vat_rate=invoice.get("vat_rate", 0.20),
        discount_total=invoice.get("discount_total", 0),
        retention_percent=invoice.get("retention_percent", 0),
        advance_amount=invoice.get("advance_amount", 0),
        paid_total=invoice.get("paid_total", 0),
        currency=invoice.get("currency", "EUR"),
    )

    issue_date = invoice.get("issue_date")
    tax_event_date = invoice.get("tax_event_date")
    due_date = invoice.get("due_date")
    period_from = invoice.get("period_from")
    period_to = invoice.get("period_to")

    updates = {
        "xl/worksheets/sheet1.xml": {
            "C1": display["name"],
            "C4": display["phone_email"],
            "J3": invoice.get("invoice_number", ""),
            "J4": invoice.get("currency", "EUR"),
            "A8": format_date(issue_date),
            "D8": format_date(tax_event_date),
            "G8": format_date(due_date),
            "I8": _short_text(_translated_payment_method(invoice.get("payment_method", ""), language)),
            "A11": _short_text(invoice.get("project_name", "")),
            "D11": _short_text(invoice.get("site_address", "")),
            "G11": _short_text(invoice.get("contract_no", "")),
            "I11": _short_text(invoice.get("protocol_no", "")),
            "A13": format_date(period_from) if period_from else "",
            "D13": format_date(period_to) if period_to else "",
            "G13": _short_text(invoice.get("order_reference", "")),
            "I13": _short_text(invoice.get("issue_place", "")),
            "C16": display["name"],
            "C17": display["eik"],
            "C18": display["vat"],
            "C19": display["address"],
            "C20": display["director"],
            "C21": display["phone_email"],
            "I16": _short_text(invoice.get("customer_name", "")),
            "I17": _short_text(invoice.get("customer_eik", "")),
            "I18": _short_text(invoice.get("customer_vat", "")),
            "I19": _short_text(invoice.get("customer_address", "")),
            "I20": _short_text(invoice.get("customer_contact", "")),
            "I21": customer_phone_email,
            "A45": _short_text(invoice.get("note", "") or text["default_note"], 180),
            "J44": totals["subtotal"],
            "J45": totals["discount_total"],
            "J46": totals["tax_base"],
            "J47": totals["vat_total"],
            "J48": totals["gross_total"],
            "J49": totals["retention_percent"],
            "K49": totals["retention_amount"],
            "J50": totals["advance_amount"],
            "J51": totals["paid_total"],
            "J52": totals["balance_total"],
            "J53": totals["balance_total"] * Decimal(str(invoice.get("exchange_rate", 1.95583))) if invoice.get("currency", "EUR") == "EUR" else "",
            "J54": format_date(due_date),
            "A51": f"{text['total']}: {format_currency(totals['balance_total'] if totals['balance_total'] else totals['gross_total'], invoice.get('currency', 'EUR'))}",
            "A58": display["bank_line"],
            "D58": display["iban"],
            "H58": display["bic"],
            "A60": text["payment_reference"].format(number=invoice.get("invoice_number", "")),
        }
    }
    vat_caption = text["vat"].format(vat=f"{_number_value(invoice.get('vat_rate', 0.20)) * 100:g}")
    updates["xl/worksheets/sheet1.xml"].update(
        {
            "H1": text["tax_document"], "H2": text["invoice"], "C3": text["construction"], "H4": text["currency"],
            "A7": text["issue_date"], "D7": text["tax_event"], "G7": text["due_date"], "I7": text["payment_method"],
            "A10": text["project"], "D10": text["site"], "G10": text["contract"], "I10": text["protocol"],
            "A12": text["period_from"], "D12": text["period_to"], "G12": text["reference"], "I12": text["issue_place"],
            "A15": text["supplier"], "G15": text["customer"], "A16": text["company"], "G16": text["company"],
            "A17": text["eik"], "G17": text["eik"], "A18": text["vat_no"], "G18": text["vat_no"],
            "A19": text["address"], "G19": text["address"], "A20": text["manager"], "G20": text["manager"],
            "A21": text["contact"], "G21": text["contact"], "A23": text["breakdown"], "B24": text["category"],
            "C24": text["description"], "D24": text["unit"], "E24": text["quantity"], "F24": text["unit_price"],
            "G24": text["discount"], "H24": text["net"], "I24": vat_caption, "J24": text["gross"], "K24": text["code"],
            "A44": text["notes"], "G44": text["subtotal"], "G45": text["adjustment"], "G46": text["tax_base"],
            "G47": vat_caption, "G48": text["total"], "G49": text["retention"], "A50": text["words"], "G50": text["advance"],
            "G51": text["paid"], "G52": text["balance"], "G53": text["reference_bgn"], "G54": text["payment_deadline"],
            "A56": text["bank_details"], "A57": text["bank"], "G60": text["bank_hint"],
            "A62": f"{text['prepared']} ________________________________________________", "G62": f"{text['received']} ________________________________________________",
        }
    )

    items = invoice.get("items", [])
    for idx in range(MAX_VISIBLE_ITEMS):
        row = 25 + idx
        if idx < len(items):
            item = items[idx]
            line = {
                "A": idx + 1,
                "B": _short_text(_translated_category(item.get("category", ""), language)),
                "C": _short_text(_translated_standard_description(item.get("description", ""), language, invoice.get("invoice_kind")), 90),
                "D": _short_text(_translated_unit(item.get("unit", ""), language)),
                "E": _number_value(item.get("quantity")),
                "F": _number_value(item.get("unit_price")),
                "G": _number_value(item.get("discount_percent")) / 100.0 if _number_value(item.get("discount_percent")) > 1 else _number_value(item.get("discount_percent")),
                "H": _number_value(item.get("net_amount")),
                "I": _number_value(item.get("vat_amount")),
                "J": _number_value(item.get("gross_amount")),
                "K": _short_text(item.get("code_stage", "")),
            }
        else:
            line = {"A": "", "B": "", "C": "", "D": "", "E": "", "F": "", "G": "", "H": "", "I": "", "J": "", "K": ""}
        for col, value in line.items():
            # Clear pre-existing template formulas for unused rows. Otherwise Excel
            # can surface formula errors on rows that the user did not enter.
            updates["xl/worksheets/sheet1.xml"][f"{col}{row}"] = value

    return updates


def _export_invoice_xlsx_via_excel(template_path: Path, output_path: Path, updates: dict[str, dict[str, Any]]) -> Path:
    """Fill a copied workbook through Excel so every saved copy stays Excel-valid."""
    sheet_updates = updates.get("xl/worksheets/sheet1.xml", {})
    encoded_updates = base64.b64encode(
        json.dumps(sheet_updates, ensure_ascii=False, default=str).encode("utf-8")
    ).decode("ascii")
    template_path = Path(template_path).resolve()
    output_path = Path(output_path).resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    script = f"""
$ErrorActionPreference = 'Stop'
$templatePath = {_powershell_literal(template_path)}
$outputPath = {_powershell_literal(output_path)}
$encodedUpdates = {json.dumps(encoded_updates)}
$excel = $null
$book = $null
$sheet = $null
try {{
    Copy-Item -LiteralPath $templatePath -Destination $outputPath -Force
    $json = [System.Text.Encoding]::UTF8.GetString([System.Convert]::FromBase64String($encodedUpdates))
    $updates = ConvertFrom-Json -InputObject $json
    $excel = New-Object -ComObject Excel.Application
    $excel.Visible = $false
    $excel.DisplayAlerts = $false
    $excel.ScreenUpdating = $false
    $excel.EnableEvents = $false
    $book = $excel.Workbooks.Open($outputPath)
    $sheet = $book.Worksheets.Item(1)
    # Every invoice is a single A4 portrait document.  Do this on every saved
    # workbook rather than relying on a user-edited template's print settings.
    $sheet.PageSetup.Orientation = 1 # xlPortrait
    $sheet.PageSetup.PaperSize = 9 # xlPaperA4
    $sheet.PageSetup.Zoom = $false
    $sheet.PageSetup.FitToPagesWide = 1
    $sheet.PageSetup.FitToPagesTall = 1
    $sheet.PageSetup.LeftMargin = $excel.CentimetersToPoints(0.45)
    $sheet.PageSetup.RightMargin = $excel.CentimetersToPoints(0.45)
    $sheet.PageSetup.TopMargin = $excel.CentimetersToPoints(0.45)
    $sheet.PageSetup.BottomMargin = $excel.CentimetersToPoints(0.45)
    $sheet.PageSetup.HeaderMargin = $excel.CentimetersToPoints(0.15)
    $sheet.PageSetup.FooterMargin = $excel.CentimetersToPoints(0.15)
    $sheet.PageSetup.CenterHorizontally = $true
    $sheet.PageSetup.CenterVertically = $false
    $sheet.PageSetup.PrintArea = '{INVOICE_PRINT_AREA}'
    foreach ($property in $updates.PSObject.Properties) {{
        $range = $sheet.Range($property.Name)
        if ($null -eq $property.Value) {{
            $range.ClearContents()
        }} elseif ($property.Value -is [int] -or $property.Value -is [long] -or $property.Value -is [double] -or $property.Value -is [decimal]) {{
            # Excel COM accepts a Double reliably, including integer quantity fields.
            $range.Value2 = [double]$property.Value
        }} else {{
            $range.Value2 = [string]$property.Value
        }}
    }}
    $book.Save()
}} finally {{
    if ($book) {{ $book.Close($true) }}
    if ($sheet) {{ [void][System.Runtime.InteropServices.Marshal]::FinalReleaseComObject($sheet) }}
    if ($book) {{ [void][System.Runtime.InteropServices.Marshal]::FinalReleaseComObject($book) }}
    if ($excel) {{ $excel.Quit(); [void][System.Runtime.InteropServices.Marshal]::FinalReleaseComObject($excel) }}
    [GC]::Collect()
    [GC]::WaitForPendingFinalizers()
}}
"""
    try:
        result = subprocess.run(
            [_powershell_executable(), "-NoProfile", "-NonInteractive", "-WindowStyle", "Hidden", "-Command", script],
            check=False,
            capture_output=True,
            text=True,
            # A protected customer template can take longer than one minute on
            # a busy workstation.  This task runs in the export worker, so a
            # generous limit never freezes the invoice editor.
            timeout=180,
        )
    except (OSError, subprocess.TimeoutExpired) as exc:
        raise RuntimeError(
            "Excel šablon nije završio pripremu u predviđenom vremenu. "
            "Sačekajte da se zatvore eventualno otvoreni Excel prozori i pokušajte ponovo."
        ) from exc
    if result.returncode != 0 or not output_path.exists() or output_path.stat().st_size == 0:
        details = "\n".join(part for part in [result.stdout.strip(), result.stderr.strip()] if part)
        raise RuntimeError(f"Excel nije uspeo da popuni kopiju originalnog šablona.\n{details}".strip())
    return output_path


def export_invoice_xlsx(invoice: dict[str, Any], output_path: Path, template_path: Path | None = None) -> Path:
    template = template_path or TEMPLATE_XLSX
    if not template.exists():
        raise FileNotFoundError(f"Template workbook not found: {template}")
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    updates = build_invoice_xlsx_updates(invoice)
    # A custom workbook can contain Excel-only relationships that survive a
    # ZIP/XML edit but make the resulting file impossible for Excel COM to
    # reopen.  Use Excel itself on Windows so the editable copy and its A4 PDF
    # are always generated from the exact original template.  This runs in the
    # background export worker and is therefore safe even for larger files.
    if sys.platform.startswith("win"):
        return _export_invoice_xlsx_via_excel(template, output_path, updates)
    _copy_zip_with_updates(template, output_path, updates)
    return output_path


def _pdf_font(name_regular: str = FONT_REGULAR, name_bold: str = FONT_BOLD) -> tuple[str, str]:
    return name_regular, name_bold


def _pstyle(name: str, size: int, bold: bool = False, color=TEXT_DARK, align=TA_LEFT, leading: int | None = None) -> ParagraphStyle:
    regular, bold_name = _pdf_font()
    return ParagraphStyle(
        name=name,
        fontName=bold_name if bold else regular,
        fontSize=size,
        leading=leading or int(size * 1.2),
        textColor=color,
        alignment=align,
    )


def _draw_paragraph(c: canvas.Canvas, text: str, x: float, y: float, w: float, h: float, style: ParagraphStyle) -> float:
    para = Paragraph(text, style)
    aw, ah = para.wrap(w, h)
    para.drawOn(c, x, y + h - ah)
    return ah


def _draw_box(c: canvas.Canvas, x: float, y: float, w: float, h: float, fill=None, stroke=None, radius: float = 0) -> None:
    if fill:
        c.setFillColor(fill)
    if stroke:
        c.setStrokeColor(stroke)
    c.rect(x, y, w, h, stroke=1 if stroke else 0, fill=1 if fill else 0)


def _draw_header(c: canvas.Canvas, invoice: dict[str, Any], company: dict[str, Any], logo_path: Path | None) -> None:
    c.setFillColor(GREEN_DARK)
    c.rect(0, PAGE_H - 20 * mm, PAGE_W, 20 * mm, stroke=0, fill=1)
    if logo_path and logo_path.exists():
        try:
            c.drawImage(ImageReader(str(logo_path)), MARGIN_X, PAGE_H - 40 * mm, width=28 * mm, height=28 * mm, mask="auto", preserveAspectRatio=True, anchor="nw")
        except Exception:
            pass

    regular, bold = _pdf_font()
    c.setFont(bold, 15)
    c.setFillColor(GREEN_DARK)
    c.drawString(MARGIN_X + 32 * mm, PAGE_H - 24 * mm, _short_text(company.get("name", "OpsNest"), 42))
    c.setFont(regular, 8.5)
    c.setFillColor(TEXT_MUTED)
    top_line = " • ".join(filter(None, [_short_text(company.get("phone", ""), 26), _short_text(company.get("email", ""), 32)]))
    c.drawString(MARGIN_X + 32 * mm, PAGE_H - 30 * mm, top_line)
    c.setFillColor(colors.white)
    c.setFont(bold, 9)
    c.drawRightString(PAGE_W - MARGIN_X, PAGE_H - 9.5 * mm, "ДАНЪЧЕН ДОКУМЕНТ")
    c.setFont(bold, 17)
    c.drawRightString(PAGE_W - MARGIN_X, PAGE_H - 18.5 * mm, "ФАКТУРА")
    c.setFont(bold, 9)
    c.drawRightString(PAGE_W - MARGIN_X, PAGE_H - 26.5 * mm, f"№ {invoice.get('invoice_number', '')}")
    c.setFont(regular, 8.5)
    c.drawRightString(PAGE_W - MARGIN_X, PAGE_H - 32 * mm, invoice.get("currency", "EUR"))


def _draw_key_value_row(c: canvas.Canvas, x: float, y_top: float, labels: list[str], values: list[str], widths: list[float], row_h: float = 17) -> float:
    regular, bold = _pdf_font()
    y = y_top - row_h
    cursor = x
    for idx, (label, value, w) in enumerate(zip(labels, values, widths)):
        _draw_box(c, cursor, y, w, row_h, fill=GREEN_LIGHT if idx % 2 == 0 else GREEN_LIGHTER, stroke=GREEN_LINE)
        c.setFillColor(TEXT_MUTED)
        c.setFont(bold, 7.8)
        c.drawString(cursor + 2.5 * mm, y + row_h - 6.5, label)
        c.setFillColor(TEXT_DARK)
        c.setFont(regular, 8.2)
        c.drawString(cursor + 2.5 * mm, y + 4.5, _short_text(value, int(w / 2.7)))
        cursor += w
    return y


def _draw_block(c: canvas.Canvas, x: float, y_top: float, w: float, title: str, labels: list[str], values: list[str], row_h: float = 15.5) -> None:
    regular, bold = _pdf_font()
    header_h = 16
    _draw_box(c, x, y_top - header_h, w, header_h, fill=GREEN_DARK, stroke=GREEN_DARK)
    c.setFillColor(colors.white)
    c.setFont(bold, 8.6)
    c.drawString(x + 2.2 * mm, y_top - 11.5, title)
    body_y = y_top - header_h
    half = w / 2
    for idx, (label, value) in enumerate(zip(labels, values)):
        y = body_y - (idx + 1) * row_h
        _draw_box(c, x, y, half, row_h, fill=GREEN_LIGHT if idx % 2 == 0 else GREEN_LIGHTER, stroke=GREEN_LINE)
        _draw_box(c, x + half, y, half, row_h, fill=GREEN_LIGHT if idx % 2 == 0 else GREEN_LIGHTER, stroke=GREEN_LINE)
        c.setFillColor(TEXT_MUTED)
        c.setFont(bold, 7.4)
        c.drawString(x + 2.2 * mm, y + row_h - 5.8, label)
        c.setFillColor(TEXT_DARK)
        c.setFont(regular, 8.0)
        c.drawString(x + half + 2.2 * mm, y + 4.0, _short_text(value, int((half - 6 * mm) / 2.6)))


def _draw_items_table(c: canvas.Canvas, invoice: dict[str, Any], x: float, y_top: float, w: float) -> float:
    styles = getSampleStyleSheet()
    item_font = 6.4
    data = [[
        "№",
        "Категория",
        "Вид СМР / описание",
        "Мярка",
        "Количество",
        "Ед. цена без ДДС",
        "Отстъпка %",
        "Стойност без ДДС",
        "ДДС 20%",
        "Общо с ДДС",
        "Код / етап",
    ]]
    for idx in range(MAX_VISIBLE_ITEMS):
        if idx < len(invoice.get("items", [])):
            item = invoice["items"][idx]
            data.append(
                [
                    str(idx + 1),
                    _short_text(item.get("category", ""), 14),
                    _short_text(item.get("description", ""), 54),
                    _short_text(item.get("unit", ""), 8),
                    f"{_number_value(item.get('quantity')):,.2f}".replace(",", "."),
                    f"{_number_value(item.get('unit_price')):,.2f}".replace(",", "."),
                    f"{_number_value(item.get('discount_percent')):,.2f}".replace(",", "."),
                    f"{_number_value(item.get('net_amount')):,.2f}".replace(",", "."),
                    f"{_number_value(item.get('vat_amount')):,.2f}".replace(",", "."),
                    f"{_number_value(item.get('gross_amount')):,.2f}".replace(",", "."),
                    _short_text(item.get("code_stage", ""), 10),
                ]
            )
        else:
            data.append(["", "", "", "", "", "", "", "", "", "", ""])
    col_widths = [
        8 * mm,
        13 * mm,
        48 * mm,
        9 * mm,
        13 * mm,
        18 * mm,
        10 * mm,
        18 * mm,
        10 * mm,
        18 * mm,
        15 * mm,
    ]
    table = Table(data, colWidths=col_widths, rowHeights=[6.2 * mm] + [4.4 * mm] * MAX_VISIBLE_ITEMS)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), GREEN_DARK),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), FONT_BOLD),
                ("FONTNAME", (0, 1), (-1, -1), FONT_REGULAR),
                ("FONTSIZE", (0, 0), (-1, -1), item_font),
                ("LEADING", (0, 0), (-1, -1), 7),
                ("ALIGN", (0, 0), (0, -1), "CENTER"),
                ("ALIGN", (4, 1), (9, -1), "RIGHT"),
                ("ALIGN", (1, 1), (3, -1), "LEFT"),
                ("ALIGN", (10, 1), (10, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("BACKGROUND", (0, 1), (-1, -1), GREEN_LIGHTER),
                ("GRID", (0, 0), (-1, -1), 0.45, GREEN_LINE),
                ("LEFTPADDING", (0, 0), (-1, -1), 2),
                ("RIGHTPADDING", (0, 0), (-1, -1), 2),
                ("TOPPADDING", (0, 0), (-1, -1), 1),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
            ]
        )
    )
    tw, th = table.wrapOn(c, w, 0)
    table.drawOn(c, x, y_top - th)
    return y_top - th


def _draw_summary_panel(c: canvas.Canvas, invoice: dict[str, Any], x: float, y_top: float, w: float) -> float:
    regular, bold = _pdf_font()
    totals = calculate_invoice_totals(
        invoice.get("items", []),
        vat_rate=invoice.get("vat_rate", 0.20),
        discount_total=invoice.get("discount_total", 0),
        retention_percent=invoice.get("retention_percent", 0),
        advance_amount=invoice.get("advance_amount", 0),
        paid_total=invoice.get("paid_total", 0),
        currency=invoice.get("currency", "EUR"),
    )
    rows = [
        ("СТОЙНОСТ НА СМР БЕЗ ДДС", format_currency(totals["subtotal"], invoice.get("currency", "EUR"))),
        ("ДОПЪЛНИТЕЛНА ОТСТЪПКА / КОРЕКЦИЯ БЕЗ ДДС", format_currency(totals["discount_total"], invoice.get("currency", "EUR"))),
        ("ДАНЪЧНА ОСНОВА", format_currency(totals["tax_base"], invoice.get("currency", "EUR"))),
        ("ДДС 20%", format_currency(totals["vat_total"], invoice.get("currency", "EUR"))),
        ("ОБЩО С ДДС", format_currency(totals["gross_total"], invoice.get("currency", "EUR"))),
        ("ГАРАНЦИОННО УДЪРЖАНЕ ПО ДОГОВОР", f"{_number_value(invoice.get('retention_percent', 0)) * 100:.2f}%"),
        ("ПРИСПАДНАТ АВАНС (С ДДС)", format_currency(totals["advance_amount"], invoice.get("currency", "EUR"))),
        ("ПЛАТЕНО", format_currency(totals["paid_total"], invoice.get("currency", "EUR"))),
        ("ОСТАТЪК ЗА ПЛАЩАНЕ", format_currency(totals["balance_total"], invoice.get("currency", "EUR"))),
    ]
    row_h = 4.0 * mm
    box_h = row_h * len(rows)
    y = y_top - box_h
    for idx, (label, value) in enumerate(rows):
        row_y = y + box_h - (idx + 1) * row_h
        fill = GREEN_DARK if idx in {4, 8} else GREEN_LIGHT if idx % 2 == 0 else GREEN_LIGHTER
        text_color = colors.white if idx in {4, 8} else TEXT_MUTED
        value_color = colors.white if idx in {4, 8} else TEXT_DARK
        _draw_box(c, x, row_y, w, row_h, fill=fill, stroke=GREEN_LINE)
        c.setFont(bold if idx not in {4, 8} else FONT_BOLD, 6.8 if idx not in {4, 8} else 7.6)
        c.setFillColor(text_color)
        c.drawRightString(x + w - 3 * mm, row_y + row_h - 5.2, label)
        if value:
            c.setFillColor(value_color)
            c.setFont(regular, 7.0)
            if idx in {4, 8}:
                c.setFont(bold, 7.6)
            c.drawRightString(x + w - 3 * mm, row_y + 1.0, value)
    return y


def _draw_note_box(c: canvas.Canvas, x: float, y_top: float, w: float, h: float, invoice: dict[str, Any]) -> None:
    regular, bold = _pdf_font()
    _draw_box(c, x, y_top - h, w, h, fill=GREEN_LIGHTER, stroke=GREEN_LINE)
    header_h = 11
    _draw_box(c, x, y_top - header_h, w, header_h, fill=GREEN_LIGHT, stroke=GREEN_LINE)
    c.setFillColor(GREEN_DARK)
    c.setFont(bold, 8.0)
    c.drawString(x + 2.2 * mm, y_top - 8.5, "ОСНОВАНИЕ И БЕЛЕЖКИ")
    text = invoice.get("note") or (
        "Извършени строително-монтажни работи съгласно договор, количествено-стойностна сметка и/или подписан протокол. "
        "Попълнете допълнителни условия, етап, гаранционни удръжки и срокове."
    )
    style = _pstyle("note", 7.3, bold=False, color=TEXT_DARK, align=TA_LEFT, leading=9)
    para = Paragraph(_short_text(text, 380), style)
    aw, ah = para.wrap(w - 4 * mm, h - header_h - 4 * mm)
    para.drawOn(c, x + 2 * mm, y_top - h + 2 * mm)


def _draw_bank_block(c: canvas.Canvas, invoice: dict[str, Any], x: float, y_top: float, w: float) -> None:
    company = invoice.get("company", {})
    regular, bold = _pdf_font()
    row_h = 8.0
    _draw_box(c, x, y_top - row_h * 2, w, row_h * 2, fill=GREEN_LIGHTER, stroke=GREEN_LINE)
    _draw_box(c, x, y_top - row_h, w, row_h, fill=GREEN_DARK, stroke=GREEN_DARK)
    labels = [("БАНКА", company.get("bank_name", "")), ("IBAN", company.get("iban", "")), ("BIC / SWIFT", company.get("bic", ""))]
    col_w = w / 3
    for idx, (label, value) in enumerate(labels):
        xx = x + idx * col_w
        _draw_box(c, xx, y_top - row_h * 2, col_w, row_h, fill=GREEN_LIGHTER, stroke=GREEN_LINE)
        c.setFillColor(TEXT_MUTED)
        c.setFont(bold, 7.5)
        c.drawString(xx + 2 * mm, y_top - row_h - 6.8, label)
        c.setFillColor(TEXT_DARK)
        c.setFont(regular, 7.9)
        c.drawString(xx + 2 * mm, y_top - row_h * 2 + 3.0, _short_text(value, int((col_w - 4 * mm) / 2.6)))
    c.setFillColor(colors.white)
    c.setFont(bold, 7.9)
    c.drawString(x + 2.2 * mm, y_top - 6.0, "БАНКОВИ ДАННИ ЗА ПЛАЩАНЕ")


def _resolve_node_executable() -> Path:
    env = os.environ.get("DELTA_FAKTURE_NODE")
    if env:
        candidate = Path(env)
        if candidate.exists():
            return candidate
    for candidate in [LOCAL_RUNTIME_NODE_EXE, CODEx_RUNTIME_NODE_EXE]:
        if candidate.exists():
            return candidate
    which = shutil.which("node")
    if which:
        candidate = Path(which)
        try:
            result = subprocess.run([str(candidate), "--version"], check=False, capture_output=True, text=True, timeout=10)
            if result.returncode == 0:
                version = result.stdout.strip().lstrip("v")
                major = int(version.split(".")[0]) if version else 0
                if major >= 18:
                    return candidate
        except Exception:
            pass
    raise FileNotFoundError("Node.js executable not found for workbook rendering.")


def _resolve_artifact_tool_module() -> Path:
    env = os.environ.get("DELTA_ARTIFACT_TOOL")
    if env:
        candidate = Path(env)
        if candidate.exists():
            return candidate
    for candidate in [LOCAL_ARTIFACT_TOOL, CODEx_ARTIFACT_TOOL]:
        if candidate.exists():
            return candidate
    raise FileNotFoundError("artifact-tool module not found for workbook rendering.")


def _render_xlsx_to_png(xlsx_path: Path, png_path: Path, sheet_name: str = "Фактура") -> Path:
    node_exe = _resolve_node_executable()
    artifact_tool = _resolve_artifact_tool_module().resolve().as_uri()
    xlsx_path = Path(xlsx_path).resolve()
    png_path = Path(png_path).resolve()
    png_path.parent.mkdir(parents=True, exist_ok=True)
    script = f"""
import fs from "node:fs/promises";
import {{ FileBlob, SpreadsheetFile }} from {json.dumps(artifact_tool)};
const input = await FileBlob.load({json.dumps(str(xlsx_path))});
const workbook = await SpreadsheetFile.importXlsx(input);
const png = await workbook.render({{
  sheetName: {json.dumps(sheet_name)},
  autoCrop: "all",
  // This is the no-Excel fallback; 2x keeps the invoice legible without a long wait.
  scale: 2,
  format: "png",
}});
await fs.writeFile({json.dumps(str(png_path))}, new Uint8Array(await png.arrayBuffer()));
"""
    result = subprocess.run(
        [str(node_exe), "--input-type=module", "--eval", script],
        check=False,
        capture_output=True,
        text=True,
    )
    if result.returncode != 0:
        raise RuntimeError(
            "Failed to render workbook to PNG.\n"
            f"STDOUT:\n{result.stdout}\nSTDERR:\n{result.stderr}"
        )
    return png_path


def _export_xlsx_to_pdf_via_excel(xlsx_path: Path, output_path: Path) -> bool:
    """Use the invoice sheet's native Excel print engine when it is available on Windows."""
    if not sys.platform.startswith("win"):
        return False
    xlsx_path = Path(xlsx_path).resolve()
    output_path = Path(output_path).resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    script = f"""
$ErrorActionPreference = 'Stop'
$excel = $null
$book = $null
$sheet = $null
try {{
    $excel = New-Object -ComObject Excel.Application
    $excel.Visible = $false
    $excel.DisplayAlerts = $false
    $book = $excel.Workbooks.Open({_powershell_literal(xlsx_path)}, 0, $true)
    # The workbook includes support sheets. Export only the first, protected invoice sheet.
    $sheet = $book.Worksheets.Item(1)
    # Older saved copies may predate the print rule, so enforce it immediately
    # before every PDF export as well.
    $sheet.PageSetup.Orientation = 1 # xlPortrait
    $sheet.PageSetup.PaperSize = 9 # xlPaperA4
    $sheet.PageSetup.Zoom = $false
    $sheet.PageSetup.FitToPagesWide = 1
    $sheet.PageSetup.FitToPagesTall = 1
    $sheet.PageSetup.LeftMargin = $excel.CentimetersToPoints(0.45)
    $sheet.PageSetup.RightMargin = $excel.CentimetersToPoints(0.45)
    $sheet.PageSetup.TopMargin = $excel.CentimetersToPoints(0.45)
    $sheet.PageSetup.BottomMargin = $excel.CentimetersToPoints(0.45)
    $sheet.PageSetup.HeaderMargin = $excel.CentimetersToPoints(0.15)
    $sheet.PageSetup.FooterMargin = $excel.CentimetersToPoints(0.15)
    $sheet.PageSetup.CenterHorizontally = $true
    $sheet.PageSetup.CenterVertically = $false
    $sheet.PageSetup.PrintArea = '{INVOICE_PRINT_AREA}'
    $sheet.ExportAsFixedFormat(0, {_powershell_literal(output_path)}, 0, $true, $false)
}} finally {{
    if ($book) {{ $book.Close($false) }}
    if ($sheet) {{ [void][System.Runtime.InteropServices.Marshal]::FinalReleaseComObject($sheet) }}
    if ($book) {{ [void][System.Runtime.InteropServices.Marshal]::FinalReleaseComObject($book) }}
    if ($excel) {{ $excel.Quit(); [void][System.Runtime.InteropServices.Marshal]::FinalReleaseComObject($excel) }}
    [GC]::Collect()
    [GC]::WaitForPendingFinalizers()
}}
"""
    try:
        result = subprocess.run(
            [_powershell_executable(), "-NoProfile", "-NonInteractive", "-WindowStyle", "Hidden", "-Command", script],
            check=False,
            capture_output=True,
            text=True,
            timeout=120,
        )
        return result.returncode == 0 and output_path.exists() and output_path.stat().st_size > 0
    except (OSError, subprocess.TimeoutExpired):
        return False


def export_xlsx_to_pdf(xlsx_path: Path, output_path: Path) -> Path:
    """Export the saved workbook through Excel, preserving printable text and layout."""
    xlsx_path = Path(xlsx_path)
    output_path = Path(output_path)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Excel kopija fakture nije pronađena: {xlsx_path}")
    if not _export_xlsx_to_pdf_via_excel(xlsx_path, output_path):
        raise RuntimeError("Microsoft Excel je potreban za pravi PDF iz originalnog Excel šablona.")
    return output_path


def _export_invoice_pdf_exact(invoice: dict[str, Any], output_path: Path, template_path: Path | None = None) -> Path:
    with tempfile.TemporaryDirectory(prefix="delta_invoice_") as tmpdir:
        tmpdir_path = Path(tmpdir)
        xlsx_path = tmpdir_path / "invoice.xlsx"
        png_path = tmpdir_path / "invoice.png"
        export_invoice_xlsx(invoice, xlsx_path, template_path=template_path)
        if sys.platform.startswith("win"):
            return export_xlsx_to_pdf(xlsx_path, output_path)
        _render_xlsx_to_png(xlsx_path, png_path)
        from reportlab.lib.utils import ImageReader as _ImageReader

        image = _ImageReader(str(png_path))
        width_px, height_px = image.getSize()
        # Keep the full worksheet inside the printable A4 portrait area without
        # stretching it.  This fallback is used only when native Excel is absent.
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        c = canvas.Canvas(str(output_path), pagesize=A4)
        printable_w, printable_h = PAGE_W - 10 * mm, PAGE_H - 10 * mm
        scale = min(printable_w / width_px, printable_h / height_px)
        draw_w, draw_h = width_px * scale, height_px * scale
        c.drawImage(
            image,
            (PAGE_W - draw_w) / 2,
            (PAGE_H - draw_h) / 2,
            width=draw_w,
            height=draw_h,
            mask="auto",
        )
        c.showPage()
        c.save()
    return output_path


def export_invoice_pdf(
    invoice: dict[str, Any],
    output_path: Path,
    logo_path: Path | None = None,
    template_path: Path | None = None,
) -> Path:
    try:
        return _export_invoice_pdf_exact(invoice, output_path, template_path=template_path or TEMPLATE_XLSX)
    except Exception as exc:
        # The user requires the supplied workbook layout, so never silently replace it
        # with a different PDF design when the template renderer has a problem.
        raise RuntimeError("Originalni Excel šablon fakture nije moguće prikazati kao PDF.") from exc
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    c = canvas.Canvas(str(output_path), pagesize=A4)
    company = invoice.get("company", {})
    logo = logo_path if logo_path and logo_path.exists() else (LOGO_FILE if LOGO_FILE.exists() else None)
    _draw_header(c, invoice, company, logo)

    regular, bold = _pdf_font()
    c.setFillColor(TEXT_DARK)

    y = PAGE_H - 44 * mm
    meta_labels_1 = ["ДАТА НА ИЗДАВАНЕ", "ДАТА НА ДАНЪЧНОТО СЪБИТИЕ", "ПАДЕЖ", "НАЧИН НА ПЛАЩАНЕ"]
    meta_values_1 = [format_date(invoice.get("issue_date")), format_date(invoice.get("tax_event_date")), format_date(invoice.get("due_date")), _short_text(invoice.get("payment_method", ""))]
    _draw_key_value_row(c, MARGIN_X, y, meta_labels_1, meta_values_1, [CONTENT_W / 4] * 4, row_h=16)

    y -= 22
    meta_labels_2 = ["ОБЕКТ / ПРОЕКТ", "АДРЕС НА ОБЕКТА", "ДОГОВОР №", "ПРОТОКОЛ / АКТ ОБР. 19 №"]
    meta_values_2 = [_short_text(invoice.get("project_name", ""), 28), _short_text(invoice.get("site_address", ""), 30), _short_text(invoice.get("contract_no", ""), 18), _short_text(invoice.get("protocol_no", ""), 18)]
    _draw_key_value_row(c, MARGIN_X, y, meta_labels_2, meta_values_2, [CONTENT_W / 4] * 4, row_h=16)

    y -= 22
    meta_labels_3 = ["ПЕРИОД НА СМР — ОТ", "ДО", "ПОРЪЧКА / РЕФЕРЕНЦИЯ", "МЯСТО НА ИЗДАВАНЕ"]
    meta_values_3 = [format_date(invoice.get("period_from")), format_date(invoice.get("period_to")), _short_text(invoice.get("order_reference", ""), 22), _short_text(invoice.get("issue_place", ""), 16)]
    _draw_key_value_row(c, MARGIN_X, y, meta_labels_3, meta_values_3, [CONTENT_W / 4] * 4, row_h=16)

    y -= 24
    block_w = (CONTENT_W - 10 * mm) / 2
    left_x = MARGIN_X
    right_x = MARGIN_X + block_w + 10 * mm
    supplier_labels = ["Фирма", "ЕИК / Булстат", "ИН по ДДС", "Адрес", "МОЛ / Управител", "Телефон / Имейл"]
    supplier_values = [
        _short_text(company.get("name", "")),
        _short_text(company.get("eik", "")),
        _short_text(company.get("vat_number", "")),
        _short_text(company.get("address", "")),
        _short_text(company.get("director_name", "")),
        " • ".join(filter(None, [_short_text(company.get("phone", ""), 24), _short_text(company.get("email", ""), 32)])),
    ]
    customer_labels = ["Фирма", "ЕИК / Булстат", "ИН по ДДС", "Адрес", "МОЛ", "Телефон / Имейл"]
    customer_values = [
        _short_text(invoice.get("customer_name", "")),
        _short_text(invoice.get("customer_eik", "")),
        _short_text(invoice.get("customer_vat", "")),
        _short_text(invoice.get("customer_address", "")),
        _short_text(invoice.get("customer_contact", "")),
        " • ".join(filter(None, [_short_text(invoice.get("customer_phone", ""), 24), _short_text(invoice.get("customer_email", ""), 32)])),
    ]
    _draw_block(c, left_x, y, block_w, "ИЗПЪЛНИТЕЛ / ДОСТАВЧИК", supplier_labels, supplier_values)
    _draw_block(c, right_x, y, block_w, "ВЪЗЛОЖИТЕЛ / ПОЛУЧАТЕЛ", customer_labels, customer_values)

    y -= 6 * 15.5 + 16 + 6
    table_y = y
    table_bottom = _draw_items_table(c, invoice, MARGIN_X, table_y, CONTENT_W)

    bottom_top = table_bottom - 6 * mm
    note_w = CONTENT_W * 0.56
    note_h = 33 * mm
    _draw_note_box(c, MARGIN_X, bottom_top, note_w, note_h, invoice)
    summary_x = MARGIN_X + note_w + 4 * mm
    summary_w = CONTENT_W - note_w - 4 * mm
    summary_top = bottom_top
    _draw_summary_panel(c, invoice, summary_x, summary_top, summary_w)

    words = number_to_words_bg(
        calculate_invoice_totals(
            invoice.get("items", []),
            vat_rate=invoice.get("vat_rate", 0.20),
            discount_total=invoice.get("discount_total", 0),
            retention_percent=invoice.get("retention_percent", 0),
            advance_amount=invoice.get("advance_amount", 0),
            paid_total=invoice.get("paid_total", 0),
            currency=invoice.get("currency", "EUR"),
        )["balance_total"],
        invoice.get("currency", "EUR"),
    )
    c.setFillColor(GREEN_DARK)
    c.setFont(bold, 8.4)
    c.drawString(MARGIN_X, 18 * mm, "СУМА С ДУМИ")
    c.setFont(regular, 7.5)
    c.setFillColor(TEXT_DARK)
    c.drawString(MARGIN_X, 14.5 * mm, _short_text(words, 105))
    if invoice.get("currency", "EUR") == "EUR":
        bg_amount = format_currency(
            calculate_invoice_totals(
                invoice.get("items", []),
                vat_rate=invoice.get("vat_rate", 0.20),
                discount_total=invoice.get("discount_total", 0),
                retention_percent=invoice.get("retention_percent", 0),
                advance_amount=invoice.get("advance_amount", 0),
                paid_total=invoice.get("paid_total", 0),
                currency=invoice.get("currency", "EUR"),
            )["balance_total"] * Decimal(str(invoice.get("exchange_rate", 1.95583))),
            "BGN",
        )
        c.setFont(regular, 6.8)
        c.setFillColor(TEXT_MUTED)
        c.drawString(MARGIN_X, 11.8 * mm, f"Справочно в BGN: {bg_amount}")

    bank_y = 42 * mm
    _draw_bank_block(c, invoice, MARGIN_X, bank_y + 2 * mm, CONTENT_W)

    c.setFillColor(TEXT_DARK)
    c.setFont(regular, 7.6)
    c.drawString(MARGIN_X, 8.8 * mm, "Съставил: ________________________________________________")
    c.drawRightString(PAGE_W - MARGIN_X, 8.8 * mm, "Получил / Одобрил: ________________________________________________")
    c.setFillColor(GREEN_LIGHT)
    c.rect(0, 0, PAGE_W, 7 * mm, stroke=0, fill=1)
    c.setFillColor(GREEN_DARK)
    c.setFont(regular, 6.5)
    c.drawCentredString(PAGE_W / 2, 2.6 * mm, "OpsNest invoice template")

    c.showPage()
    c.save()
    return output_path


def export_invoice_bundle(invoice: dict[str, Any], output_dir: Path, template_path: Path | None = None, logo_path: Path | None = None) -> dict[str, Path]:
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    base = safe_filename(f"faktura_{invoice.get('invoice_number', '0000000000')}")
    pdf_path = output_dir / f"{base}.pdf"
    xlsx_path = output_dir / f"{base}.xlsx"
    # Write the workbook first, then let Excel export that exact saved copy to PDF.
    export_invoice_xlsx(invoice, xlsx_path, template_path=template_path)
    if sys.platform.startswith("win"):
        export_xlsx_to_pdf(xlsx_path, pdf_path)
    else:
        export_invoice_pdf(invoice, pdf_path, logo_path=logo_path, template_path=template_path)
    return {"pdf": pdf_path, "xlsx": xlsx_path}


def _credit_note_source(note: dict[str, Any]) -> dict[str, Any]:
    source = note.get("source_invoice")
    return source if isinstance(source, dict) else {}


def export_credit_note_xlsx(note: dict[str, Any], output_path: Path) -> Path:
    """Create a standalone Excel copy for a formal credit note, never touching the invoice template."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    text = _credit_note_text(note)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = text["title"][:31]
    sheet.sheet_view.showGridLines = False
    sheet.page_setup.orientation = "portrait"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_margins.left = 0.35
    sheet.page_margins.right = 0.35
    sheet.page_margins.top = 0.45
    sheet.page_margins.bottom = 0.45
    sheet.freeze_panes = "A7"
    widths = [24, 31, 24, 31]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    dark = "245B41"
    mid = "2F7A56"
    light = "EAF4EC"
    line = "B9D5C2"
    thin = Side(style="thin", color=line)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_fill = PatternFill("solid", fgColor=dark)
    header_fill = PatternFill("solid", fgColor=mid)
    light_fill = PatternFill("solid", fgColor=light)
    title_font = Font(name="Arial", size=17, bold=True, color="FFFFFF")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    label_font = Font(name="Arial", size=9, bold=True, color=dark)
    value_font = Font(name="Arial", size=10, color="1F2937")

    sheet.merge_cells("A1:D1")
    sheet["A1"] = text["title"]
    sheet["A1"].fill = title_fill
    sheet["A1"].font = title_font
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 31
    sheet.merge_cells("A2:D2")
    sheet["A2"] = text["subtitle"]
    sheet["A2"].font = Font(name="Arial", size=10, italic=True, color="6A7A72")
    sheet["A2"].alignment = Alignment(horizontal="center")

    source = _credit_note_source(note)
    company = note.get("company") if isinstance(note.get("company"), dict) else {}
    rows = [
        (text["number"], note.get("credit_note_number") or "", text["issue_date"], format_date(note.get("issue_date"))),
        (text["source_invoice"], source.get("invoice_number") or note.get("source_invoice_number") or "", text["source_date"], format_date(source.get("issue_date"))),
        (text["project"], note.get("project_name") or source.get("project_name") or "", text["currency"], note.get("currency") or "EUR"),
    ]
    for row_no, values in enumerate(rows, start=4):
        for col_no, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_no, column=col_no, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.fill = light_fill if col_no in {1, 3} else PatternFill("solid", fgColor="FFFFFF")
            cell.font = label_font if col_no in {1, 3} else value_font
        sheet.row_dimensions[row_no].height = 24

    for title, start_col, values in [
        (text["supplier"], 1, [company.get("name", ""), company.get("eik", ""), company.get("vat_number", ""), company.get("address", "")]),
        (text["customer"], 3, [source.get("customer_name", note.get("customer_name", "")), source.get("customer_eik", ""), source.get("customer_vat", ""), source.get("customer_address", "")]),
    ]:
        sheet.merge_cells(start_row=8, start_column=start_col, end_row=8, end_column=start_col + 1)
        header = sheet.cell(row=8, column=start_col, value=title)
        header.fill = header_fill
        header.font = header_font
        header.alignment = Alignment(horizontal="center")
        for col in range(start_col, start_col + 2):
            sheet.cell(row=8, column=col).border = border
        for offset, value in enumerate(values, start=9):
            sheet.merge_cells(start_row=offset, start_column=start_col, end_row=offset, end_column=start_col + 1)
            cell = sheet.cell(row=offset, column=start_col, value=value)
            cell.font = value_font
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            for col in range(start_col, start_col + 2):
                sheet.cell(row=offset, column=col).border = border
            sheet.row_dimensions[offset].height = 22

    sheet.merge_cells("A14:D14")
    sheet["A14"] = text["reason"]
    sheet["A14"].fill = header_fill
    sheet["A14"].font = header_font
    sheet["A14"].alignment = Alignment(horizontal="center")
    sheet["A14"].border = border
    sheet.merge_cells("A15:D16")
    reason_cell = sheet["A15"]
    reason_cell.value = note.get("reason") or ""
    reason_cell.font = value_font
    reason_cell.alignment = Alignment(vertical="top", wrap_text=True)
    reason_cell.border = border
    for row in range(15, 17):
        for col in range(1, 5):
            sheet.cell(row=row, column=col).border = border
    sheet.row_dimensions[15].height = 28
    sheet.row_dimensions[16].height = 28

    for col, label in enumerate((text["net"], text["vat"], text["total"], text["linked_invoice"]), start=1):
        cell = sheet.cell(row=18, column=col, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    values = [
        float(note.get("net_amount") or 0),
        float(note.get("vat_amount") or 0),
        float(note.get("gross_amount") or 0),
        source.get("invoice_number") or note.get("source_invoice_number") or "",
    ]
    for col, value in enumerate(values, start=1):
        cell = sheet.cell(row=19, column=col, value=value)
        cell.border = border
        cell.font = Font(name="Arial", size=11, bold=col == 3, color=dark)
        cell.fill = light_fill if col == 3 else PatternFill("solid", fgColor="FFFFFF")
        cell.alignment = Alignment(horizontal="right" if col < 4 else "center", vertical="center")
        if col < 4:
            cell.number_format = '#,##0.00 "EUR"'
    sheet.row_dimensions[19].height = 26
    sheet.merge_cells("A22:D22")
    sheet["A22"] = text["archive_note"]
    sheet["A22"].font = Font(name="Arial", size=9, italic=True, color="6A7A72")
    sheet["A22"].alignment = Alignment(wrap_text=True)
    sheet.print_area = "A1:D22"
    workbook.save(output_path)
    return output_path


def export_credit_note_pdf(note: dict[str, Any], output_path: Path, logo_path: Path | None = None) -> Path:
    """Render a readable, editable-text PDF for a separate formal credit note."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    source = _credit_note_source(note)
    company = note.get("company") if isinstance(note.get("company"), dict) else {}
    text = _credit_note_text(note)
    party_labels = INVOICE_DOCUMENT_TEXT[_credit_note_document_language(note)]
    regular, bold = _pdf_font()
    c = canvas.Canvas(str(output_path), pagesize=A4)
    c.setTitle(str(note.get("credit_note_number") or text["title"]))
    c.setFillColor(GREEN_DARK)
    c.rect(0, PAGE_H - 29 * mm, PAGE_W, 29 * mm, stroke=0, fill=1)
    logo = logo_path if logo_path and logo_path.exists() else (LOGO_FILE if LOGO_FILE.exists() else None)
    if logo:
        try:
            image = ImageReader(str(logo))
            c.drawImage(image, MARGIN_X, PAGE_H - 23 * mm, width=32 * mm, height=17 * mm, preserveAspectRatio=True, mask="auto")
        except Exception:
            pass
    c.setFillColor(colors.white)
    c.setFont(bold, 18)
    c.drawRightString(PAGE_W - MARGIN_X, PAGE_H - 12 * mm, text["title"])
    c.setFont(regular, 8.5)
    c.drawRightString(PAGE_W - MARGIN_X, PAGE_H - 18 * mm, text["subtitle"])

    y = PAGE_H - 42 * mm
    metadata = [
        (text["number"], note.get("credit_note_number") or "-"),
        (text["issue_date"], format_date(note.get("issue_date")) or "-"),
        (text["source_invoice"], source.get("invoice_number") or note.get("source_invoice_number") or "-"),
        (text["source_date"], format_date(source.get("issue_date")) or "-"),
        (text["project"], note.get("project_name") or source.get("project_name") or "-"),
        (text["currency"], note.get("currency") or "EUR"),
    ]
    cell_w = CONTENT_W / 3
    for index, (label, value) in enumerate(metadata):
        x = MARGIN_X + (index % 3) * cell_w
        row = index // 3
        top = y - row * 20 * mm
        c.setFillColor(GREEN_LIGHT)
        c.rect(x, top - 16 * mm, cell_w - 2, 16 * mm, stroke=0, fill=1)
        c.setStrokeColor(GREEN_LINE)
        c.rect(x, top - 16 * mm, cell_w - 2, 16 * mm, stroke=1, fill=0)
        c.setFillColor(GREEN_DARK)
        c.setFont(bold, 7.2)
        c.drawString(x + 4 * mm, top - 5 * mm, label)
        c.setFillColor(TEXT_DARK)
        c.setFont(regular, 9.3)
        c.drawString(x + 4 * mm, top - 11 * mm, _short_text(value, 32))

    y -= 48 * mm
    block_w = (CONTENT_W - 8 * mm) / 2
    _draw_block(
        c, MARGIN_X, y, block_w, text["supplier"],
        [party_labels["company"], party_labels["eik"], party_labels["vat_no"], party_labels["address"]],
        [company.get("name", ""), company.get("eik", ""), company.get("vat_number", ""), company.get("address", "")],
    )
    _draw_block(
        c, MARGIN_X + block_w + 8 * mm, y, block_w, text["customer"],
        [party_labels["company"], party_labels["eik"], party_labels["vat_no"], party_labels["address"]],
        [source.get("customer_name", note.get("customer_name", "")), source.get("customer_eik", ""), source.get("customer_vat", ""), source.get("customer_address", "")],
    )

    y -= 90
    c.setFillColor(GREEN_DARK)
    c.setFont(bold, 9)
    c.drawString(MARGIN_X, y, text["reason"])
    y -= 4 * mm
    c.setStrokeColor(GREEN_LINE)
    c.setFillColor(GREEN_LIGHTER)
    c.roundRect(MARGIN_X, y - 34 * mm, CONTENT_W, 34 * mm, 3 * mm, stroke=1, fill=1)
    _draw_paragraph(c, _short_text(note.get("reason") or "-", 800), MARGIN_X + 4 * mm, y - 31 * mm, CONTENT_W - 8 * mm, 28 * mm, _pstyle("credit-note-reason", 9, color=TEXT_DARK, leading=12))

    table_y = y - 48 * mm
    table_data = [
        [text["net"], text["vat"], text["total"]],
        [
            format_currency(note.get("net_amount") or 0, note.get("currency") or "EUR"),
            format_currency(note.get("vat_amount") or 0, note.get("currency") or "EUR"),
            format_currency(note.get("gross_amount") or 0, note.get("currency") or "EUR"),
        ],
    ]
    table = Table(table_data, colWidths=[CONTENT_W / 3] * 3, rowHeights=[9 * mm, 12 * mm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), GREEN_MID),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), bold),
        ("FONTNAME", (0, 1), (-1, 1), bold),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("BACKGROUND", (0, 1), (-1, 1), GREEN_LIGHT),
        ("GRID", (0, 0), (-1, -1), 0.5, GREEN_LINE),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    table.wrapOn(c, CONTENT_W, 30 * mm)
    table.drawOn(c, MARGIN_X, table_y - 21 * mm)

    c.setFillColor(TEXT_MUTED)
    c.setFont(regular, 7.3)
    c.drawString(MARGIN_X, 21 * mm, "Dokument je vezan za izvornu fakturu i evidentirani povraćaj uplate.")
    c.drawString(MARGIN_X, 16.5 * mm, "Proverite poresko knjiženje sa knjigovođom pre predaje evidencije.")
    c.setStrokeColor(GREEN_LINE)
    c.line(MARGIN_X, 12 * mm, PAGE_W - MARGIN_X, 12 * mm)
    c.setFillColor(TEXT_DARK)
    c.setFont(regular, 8)
    c.drawString(MARGIN_X, 7 * mm, "Sastavio: __________________________________________")
    c.drawRightString(PAGE_W - MARGIN_X, 7 * mm, "Primio: __________________________________________")
    c.save()
    return output_path


def export_credit_note_bundle(note: dict[str, Any], output_dir: Path, logo_path: Path | None = None) -> dict[str, Path]:
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    base = safe_filename(f"odobrenje_{note.get('credit_note_number') or note.get('id') or '0000000000'}")
    xlsx_path = export_credit_note_xlsx(note, output_dir / f"{base}.xlsx")
    pdf_path = export_credit_note_pdf(note, output_dir / f"{base}.pdf", logo_path=logo_path)
    return {"pdf": pdf_path, "xlsx": xlsx_path}


def _vat_ledger_headers(language: Any) -> tuple[str, ...]:
    return (
        report_text("date", language),
        report_text("document_type", language),
        report_text("document_number", language),
        report_text("partner", language),
        report_text("partner_vat", language),
        report_text("description", language),
        report_text("net_base", language),
        report_text("vat", language),
        report_text("total", language),
    )


def _vat_report_base_name(report: dict[str, Any]) -> str:
    language = _report_language(report)
    start = str(report.get("period_from") or "od").replace("-", "")
    end = str(report.get("period_to") or "do").replace("-", "")
    return safe_filename(f"{report_text('vat_file', language)}_{start}_{end}_{language.upper()}")


def _vat_report_rows(sheet, rows: list[dict[str, Any]], totals: dict[str, Any], section: str, language: Any) -> None:
    dark = "245B41"
    mid = "2F7A56"
    light = "EAF4EC"
    line = "B9D5C2"
    thin = Side(style="thin", color=line)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor=mid)
    total_fill = PatternFill("solid", fgColor=light)
    header_font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    body_font = Font(name="Arial", size=9, color="1F2937")
    total_font = Font(name="Arial", size=10, bold=True, color=dark)
    widths = (13, 22, 21, 28, 20, 42, 19, 16, 19)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    for col, header in enumerate(_vat_ledger_headers(language), start=1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 28
    for row_index, row in enumerate(rows, start=2):
        values = (
            format_date(row.get("document_date")),
            _localized_document_type(row.get("document_type"), language),
            row.get("document_no") or "",
            row.get("partner_name") or "",
            row.get("partner_vat") or "",
            row.get("description") or "",
            float(row.get("net_amount") or 0),
            float(row.get("vat_amount") or 0),
            float(row.get("gross_amount") or 0),
        )
        for col, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=col, value=value)
            cell.font = body_font
            cell.border = border
            cell.alignment = Alignment(horizontal="right" if col >= 7 else "left", vertical="top", wrap_text=True)
            if col >= 7:
                cell.number_format = '#,##0.00 "EUR"'
        sheet.row_dimensions[row_index].height = 24
    total_row = max(2, len(rows) + 2)
    sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=6)
    total_label = sheet.cell(
        row=total_row,
        column=1,
        value=report_text("output_vat" if section == "output" else "input_vat", language).upper(),
    )
    total_label.fill = total_fill
    total_label.font = total_font
    total_label.alignment = Alignment(horizontal="right")
    for col in range(1, 7):
        sheet.cell(row=total_row, column=col).border = border
        sheet.cell(row=total_row, column=col).fill = total_fill
    field_prefix = "output" if section == "output" else "input"
    for col, field in enumerate(("net", "vat", "gross"), start=7):
        cell = sheet.cell(row=total_row, column=col, value=float(totals.get(f"{field_prefix}_{field}") or 0))
        cell.fill = total_fill
        cell.font = total_font
        cell.border = border
        cell.alignment = Alignment(horizontal="right")
        cell.number_format = '#,##0.00 "EUR"'
    sheet.auto_filter.ref = f"A1:I{max(1, len(rows) + 1)}"
    sheet.print_title_rows = "1:1"


def export_project_vat_evidence_xlsx(report: dict[str, Any], output_path: Path) -> Path:
    """Write a readable project VAT workbook for an accountant's review."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    language = _report_language(report)
    workbook = Workbook()
    summary = workbook.active
    summary.title = report_text("sheet_summary", language)
    summary.sheet_view.showGridLines = False
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 40
    summary.column_dimensions["C"].width = 25
    summary.column_dimensions["D"].width = 25
    dark = "245B41"
    mid = "2F7A56"
    light = "EAF4EC"
    line = "B9D5C2"
    thin = Side(style="thin", color=line)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    summary.merge_cells("A1:D1")
    title = summary["A1"]
    title.value = f"{report_text('vat_title', language)} - {report_text('accountant_file', language).upper()}"
    title.fill = PatternFill("solid", fgColor=dark)
    title.font = Font(name="Arial", size=15, bold=True, color="FFFFFF")
    title.alignment = Alignment(horizontal="center", vertical="center")
    summary.row_dimensions[1].height = 30
    project = report.get("project") if isinstance(report.get("project"), dict) else {}
    company = report.get("company") if isinstance(report.get("company"), dict) else {}
    details = (
        (report_text("company", language), company.get("name") or ""),
        (company_identifier_label(language), company.get("eik") or ""),
        (report_text("project", language), project.get("name") or ""),
        (report_text("site", language), project.get("site_address") or ""),
        (report_text("period", language), f"{format_date(report.get('period_from'))} - {format_date(report.get('period_to'))}"),
        (report_text("generated", language), str(report.get("generated_at") or "").replace("T", " ")),
    )
    for row, (label, value) in enumerate(details, start=3):
        summary.cell(row=row, column=1, value=label).font = Font(name="Arial", size=9, bold=True, color=dark)
        value_cell = summary.cell(row=row, column=2, value=value)
        value_cell.font = Font(name="Arial", size=10, color="1F2937")
        for col in (1, 2):
            summary.cell(row=row, column=col).border = border
            summary.cell(row=row, column=col).alignment = Alignment(vertical="center", wrap_text=True)
        summary.row_dimensions[row].height = 22
    totals = report.get("totals") if isinstance(report.get("totals"), dict) else {}
    headers = (report_text("description", language), report_text("net_amount", language), report_text("vat", language), report_text("total", language))
    for col, label in enumerate(headers, start=1):
        cell = summary.cell(row=11, column=col, value=label)
        cell.fill = PatternFill("solid", fgColor=mid)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
    summary_rows = (
        (report_text("output_vat", language), "output_net", "output_vat", "output_gross"),
        (report_text("input_vat", language), "input_net", "input_vat", "input_gross"),
        (report_text("vat_payable", language), None, "vat_payable", None),
    )
    for row, values in enumerate(summary_rows, start=12):
        label, net_key, vat_key, gross_key = values
        summary.cell(row=row, column=1, value=label)
        for col, key in enumerate((net_key, vat_key, gross_key), start=2):
            value = float(totals.get(key) or 0) if key else ""
            cell = summary.cell(row=row, column=col, value=value)
            if key:
                cell.number_format = '#,##0.00 "EUR"'
                cell.alignment = Alignment(horizontal="right")
            cell.border = border
        for col in range(1, 5):
            summary.cell(row=row, column=col).border = border
            summary.cell(row=row, column=col).fill = PatternFill("solid", fgColor=light if row == 14 else "FFFFFF")
            summary.cell(row=row, column=col).font = Font(name="Arial", size=10, bold=row == 14, color=dark if row == 14 else "1F2937")
    summary.merge_cells("A17:D18")
    note = summary["A17"]
    note.value = report_text("vat_working_note", language)
    note.font = Font(name="Arial", size=9, italic=True, color="6A7A72")
    note.alignment = Alignment(wrap_text=True, vertical="top")

    output_sheet = workbook.create_sheet(report_text("sheet_output", language))
    _vat_report_rows(output_sheet, list(report.get("output_rows") or []), totals, "output", language)
    input_sheet = workbook.create_sheet(report_text("sheet_input", language))
    _vat_report_rows(input_sheet, list(report.get("input_rows") or []), totals, "input", language)
    control = workbook.create_sheet(report_text("sheet_control", language))
    control.sheet_view.showGridLines = False
    control.column_dimensions["A"].width = 23
    control.column_dimensions["B"].width = 18
    control.column_dimensions["C"].width = 20
    control.column_dimensions["D"].width = 30
    control.column_dimensions["E"].width = 16
    control.merge_cells("A1:E1")
    control["A1"] = report_text("control_title", language)
    control["A1"].fill = PatternFill("solid", fgColor=dark)
    control["A1"].font = Font(name="Arial", size=13, bold=True, color="FFFFFF")
    control["A1"].alignment = Alignment(horizontal="center")
    control["A3"] = report_text("foreign_currency", language)
    control["A3"].font = Font(name="Arial", size=10, bold=True, color=dark)
    control.append([report_text("type", language), report_text("date", language), report_text("number", language), report_text("partner", language), report_text("currency", language)])
    for cell in control[4]:
        cell.fill = PatternFill("solid", fgColor=mid)
        cell.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    foreign = list(report.get("foreign_currency_rows") or [])
    for row in foreign:
        control.append((_localized_document_type(row.get("document_type"), language), format_date(row.get("document_date")), row.get("document_no"), row.get("partner_name"), row.get("currency")))
    if not foreign:
        control.append((report_text("none", language), "", "", "", ""))
    missing_start = 6 + len(foreign)
    control.cell(row=missing_start, column=1, value=report_text("missing_date", language)).font = Font(name="Arial", size=10, bold=True, color=dark)
    for col, label in enumerate((report_text("type", language), report_text("number", language), report_text("partner", language), report_text("description", language), report_text("currency", language)), start=1):
        cell = control.cell(row=missing_start + 1, column=col, value=label)
        cell.fill = PatternFill("solid", fgColor=mid)
        cell.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    missing = list(report.get("missing_date_rows") or [])
    for offset, row in enumerate(missing, start=missing_start + 2):
        control.cell(offset, 1, _localized_document_type(row.get("document_type"), language))
        control.cell(offset, 2, row.get("document_no"))
        control.cell(offset, 3, row.get("partner_name"))
        control.cell(offset, 4, row.get("description"))
        control.cell(offset, 5, row.get("currency"))
    if not missing:
        control.cell(missing_start + 2, 1, report_text("none", language))
    for row in control.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    workbook.save(output_path)
    return output_path


def _vat_pdf_cell(value: Any, style: ParagraphStyle) -> Paragraph:
    text = escape(_short_text(value, 260)).replace("\n", "<br/>")
    return Paragraph(text or "-", style)


def _vat_pdf_ledger(rows: list[dict[str, Any]], title: str, language: Any, style_token: str) -> list[Any]:
    body = _pstyle(f"vat-body-{style_token}", 6.7, color=TEXT_DARK, leading=8)
    money = _pstyle(f"vat-money-{style_token}", 6.7, color=TEXT_DARK, align=TA_RIGHT, leading=8)
    header = _pstyle(f"vat-header-{style_token}", 6.6, bold=True, color=colors.white, align=TA_CENTER, leading=7.5)
    data: list[list[Any]] = [[_vat_pdf_cell(label, header) for label in _vat_ledger_headers(language)]]
    for row in rows:
        data.append([
            _vat_pdf_cell(format_date(row.get("document_date")), body),
            _vat_pdf_cell(_localized_document_type(row.get("document_type"), language), body),
            _vat_pdf_cell(row.get("document_no"), body),
            _vat_pdf_cell(row.get("partner_name"), body),
            _vat_pdf_cell(row.get("partner_vat"), body),
            _vat_pdf_cell(row.get("description"), body),
            _vat_pdf_cell(format_currency(row.get("net_amount") or 0, "EUR"), money),
            _vat_pdf_cell(format_currency(row.get("vat_amount") or 0, "EUR"), money),
            _vat_pdf_cell(format_currency(row.get("gross_amount") or 0, "EUR"), money),
        ])
    if not rows:
        data.append([_vat_pdf_cell(report_text("empty_period", language), body), *[_vat_pdf_cell("", body) for _ in range(8)]])
    table = Table(
        data,
        # The nine columns use the complete printable width of landscape A4.
        colWidths=[20 * mm, 25 * mm, 35 * mm, 35 * mm, 29 * mm, 53 * mm, 25 * mm, 23 * mm, 26 * mm],
        repeatRows=1,
        hAlign="LEFT",
    )
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), GREEN_MID),
        ("GRID", (0, 0), (-1, -1), 0.35, GREEN_LINE),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("BACKGROUND", (0, 1), (-1, -1), colors.white),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, GREEN_LIGHTER]),
    ]))
    heading = Paragraph(title, _pstyle(f"vat-heading-{style_token}", 12, bold=True, color=GREEN_DARK))
    return [heading, Spacer(1, 4 * mm), table, Spacer(1, 7 * mm)]


def export_project_vat_evidence_pdf(report: dict[str, Any], output_path: Path) -> Path:
    """Render a multi-page, printable VAT working ledger with selectable text."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    language = _report_language(report)
    project = report.get("project") if isinstance(report.get("project"), dict) else {}
    company = report.get("company") if isinstance(report.get("company"), dict) else {}
    page_size = landscape(A4)
    document = SimpleDocTemplate(
        str(output_path),
        pagesize=page_size,
        leftMargin=13 * mm,
        rightMargin=13 * mm,
        topMargin=12 * mm,
        bottomMargin=13 * mm,
        title=f"{report_text('vat_title', language)} - {project.get('name') or ''}",
    )
    title_style = _pstyle("vat-report-title", 17, bold=True, color=GREEN_DARK)
    normal_style = _pstyle("vat-report-normal", 8.5, color=TEXT_DARK)
    small_style = _pstyle("vat-report-small", 7.5, color=TEXT_MUTED)
    story: list[Any] = [
        Paragraph(report_text("vat_title", language), title_style),
        Paragraph(report_text("vat_subtitle", language), small_style),
        Spacer(1, 5 * mm),
    ]
    details_data = [
        [report_text("company", language), company.get("name") or "", report_text("project", language), project.get("name") or ""],
        [company_identifier_label(language), company.get("eik") or "", report_text("site", language), project.get("site_address") or ""],
        [report_text("period", language), f"{format_date(report.get('period_from'))} - {format_date(report.get('period_to'))}", report_text("generated", language), str(report.get("generated_at") or "").replace("T", " ")],
    ]
    details_table = Table(details_data, colWidths=[28 * mm, 74 * mm, 28 * mm, 116 * mm])
    details_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), GREEN_LIGHT),
        ("BACKGROUND", (2, 0), (2, -1), GREEN_LIGHT),
        ("TEXTCOLOR", (0, 0), (-1, -1), TEXT_DARK),
        ("FONTNAME", (0, 0), (0, -1), FONT_BOLD),
        ("FONTNAME", (2, 0), (2, -1), FONT_BOLD),
        ("FONTNAME", (1, 0), (1, -1), FONT_REGULAR),
        ("FONTNAME", (3, 0), (3, -1), FONT_REGULAR),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.35, GREEN_LINE),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.extend([details_table, Spacer(1, 6 * mm)])
    totals = report.get("totals") if isinstance(report.get("totals"), dict) else {}
    summary_data = [
        ["", report_text("net_amount", language), report_text("vat", language), report_text("total", language)],
        [report_text("output_vat", language), format_currency(totals.get("output_net") or 0), format_currency(totals.get("output_vat") or 0), format_currency(totals.get("output_gross") or 0)],
        [report_text("input_vat", language), format_currency(totals.get("input_net") or 0), format_currency(totals.get("input_vat") or 0), format_currency(totals.get("input_gross") or 0)],
        [report_text("vat_payable", language), "", format_currency(totals.get("vat_payable") or 0), ""],
    ]
    summary_table = Table(summary_data, colWidths=[70 * mm, 55 * mm, 55 * mm, 55 * mm])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), GREEN_MID),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), FONT_BOLD),
        ("FONTNAME", (0, 1), (-1, -1), FONT_REGULAR),
        ("FONTNAME", (0, 3), (-1, 3), FONT_BOLD),
        ("BACKGROUND", (0, 3), (-1, 3), GREEN_LIGHT),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("ALIGN", (1, 0), (-1, 0), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.35, GREEN_LINE),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.extend([summary_table, Spacer(1, 8 * mm)])
    story.extend(_vat_pdf_ledger(list(report.get("output_rows") or []), report_text("output_ledger", language), language, "vat-output"))
    story.extend(_vat_pdf_ledger(list(report.get("input_rows") or []), report_text("input_ledger", language), language, "vat-input"))
    foreign = list(report.get("foreign_currency_rows") or [])
    missing = list(report.get("missing_date_rows") or [])
    if foreign or missing:
        warning = report_text("vat_control_warning", language).format(
            foreign=len(foreign), missing=len(missing), control=report_text("sheet_control", language)
        )
        story.append(Paragraph(escape(warning), _pstyle("vat-warning", 8.5, bold=True, color=colors.HexColor("#9A3412"))))

    def footer(c: canvas.Canvas, doc: Any) -> None:
        c.saveState()
        c.setStrokeColor(GREEN_LINE)
        c.line(13 * mm, 8 * mm, page_size[0] - 13 * mm, 8 * mm)
        c.setFont(FONT_REGULAR, 6.8)
        c.setFillColor(TEXT_MUTED)
        c.drawString(13 * mm, 4.8 * mm, report_text("vat_footer", language))
        c.drawRightString(page_size[0] - 13 * mm, 4.8 * mm, f"{report_text('page', language)} {doc.page}")
        c.restoreState()

    document.build(story, onFirstPage=footer, onLaterPages=footer)
    return output_path


def export_project_vat_evidence_bundle(report: dict[str, Any], output_dir: Path) -> dict[str, Path]:
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    base = _vat_report_base_name(report)
    xlsx_path = export_project_vat_evidence_xlsx(report, output_dir / f"{base}.xlsx")
    pdf_path = export_project_vat_evidence_pdf(report, output_dir / f"{base}.pdf")
    return {"pdf": pdf_path, "xlsx": xlsx_path}


def _accountant_payment_headers(language: Any) -> tuple[str, ...]:
    return (
        report_text("date", language),
        report_text("type", language),
        report_text("invoice", language),
        report_text("customer", language),
        report_text("payment_method_note", language),
        report_text("total", language),
    )


def _accountant_report_base_name(report: dict[str, Any]) -> str:
    language = _report_language(report)
    project = report.get("project") if isinstance(report.get("project"), dict) else {}
    start = str(report.get("period_from") or "period")
    end = str(report.get("period_to") or "period")
    return safe_filename(f"{report_text('accountant_file', language)}_{project.get('name') or report_text('project', language)}_{start}_{end}_{language.upper()}")


def _accountant_sheet_style(sheet: Any, headers: tuple[str, ...], widths: tuple[float, ...]) -> Border:
    mid = "2F7A56"
    thin = Side(style="thin", color="B9D5C2")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.fill = PatternFill("solid", fgColor=mid)
        cell.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 28
    return border


def _accountant_summary_sheet(workbook: Workbook, report: dict[str, Any]) -> None:
    language = _report_language(report)
    sheet = workbook.create_sheet(report_text("sheet_summary", language), 0)
    sheet.sheet_view.showGridLines = False
    for column, width in {"A": 28, "B": 25, "C": 25, "D": 25}.items():
        sheet.column_dimensions[column].width = width
    dark, mid, light = "245B41", "2F7A56", "EAF4EC"
    thin = Side(style="thin", color="B9D5C2")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    sheet.merge_cells("A1:D1")
    sheet["A1"] = report_text("accountant_title", language)
    sheet["A1"].fill = PatternFill("solid", fgColor=dark)
    sheet["A1"].font = Font(name="Arial", size=15, bold=True, color="FFFFFF")
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 30
    project = report.get("project") if isinstance(report.get("project"), dict) else {}
    company = report.get("company") if isinstance(report.get("company"), dict) else {}
    details = (
        (report_text("company", language), company.get("name") or ""),
        (report_text("project", language), project.get("name") or ""),
        (report_text("site", language), project.get("site_address") or ""),
        (report_text("period", language), f"{format_date(report.get('period_from'))} - {format_date(report.get('period_to'))}"),
        (report_text("generated", language), str(report.get("generated_at") or "").replace("T", " ")),
    )
    for row, (label, value) in enumerate(details, start=3):
        sheet.cell(row=row, column=1, value=label).font = Font(name="Arial", size=9, bold=True, color=dark)
        sheet.cell(row=row, column=2, value=value).font = Font(name="Arial", size=10, color="1F2937")
        for col in (1, 2):
            sheet.cell(row=row, column=col).border = border
            sheet.cell(row=row, column=col).alignment = Alignment(vertical="center", wrap_text=True)
    headers = (report_text("summary", language), report_text("net_amount", language), report_text("vat", language), report_text("total", language))
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=10, column=col, value=header)
        cell.fill = PatternFill("solid", fgColor=mid)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
    totals = report.get("totals") if isinstance(report.get("totals"), dict) else {}
    rows = (
        (report_text("outgoing_invoices", language), totals.get("output_net"), totals.get("output_vat"), totals.get("output_gross")),
        (report_text("incoming_bills", language), totals.get("input_net"), totals.get("input_vat"), totals.get("input_gross")),
        (report_text("payments", language), totals.get("payment_total"), "", totals.get("net_collected")),
        (report_text("refunds", language), totals.get("refund_total"), "", ""),
        (report_text("vat_payable", language), "", totals.get("vat_payable"), ""),
    )
    for row_index, values in enumerate(rows, start=11):
        for col, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=col, value=float(value) if col > 1 and value not in {"", None} else value)
            cell.border = border
            cell.fill = PatternFill("solid", fgColor=light if row_index in {13, 15} else "FFFFFF")
            cell.font = Font(name="Arial", size=10, bold=row_index in {13, 15}, color=dark if row_index in {13, 15} else "1F2937")
            cell.alignment = Alignment(horizontal="right" if col > 1 else "left", vertical="center")
            if col > 1 and value not in {"", None}:
                cell.number_format = '#,##0.00 "EUR"'
    sheet.merge_cells("A18:D19")
    sheet["A18"] = report_text("accountant_note", language)
    sheet["A18"].font = Font(name="Arial", size=9, italic=True, color="6A7A72")
    sheet["A18"].alignment = Alignment(wrap_text=True, vertical="top")


def _accountant_payment_sheet(workbook: Workbook, report: dict[str, Any]) -> None:
    language = _report_language(report)
    sheet = workbook.create_sheet(report_text("sheet_payments", language))
    border = _accountant_sheet_style(sheet, _accountant_payment_headers(language), (13, 20, 20, 28, 48, 18))
    rows = list(report.get("payment_rows") or [])
    for row_index, row in enumerate(rows, start=2):
        values = (
            format_date(row.get("payment_date")),
            _localized_document_type(row.get("type"), language),
            row.get("invoice_number") or "",
            row.get("partner_name") or "",
            " | ".join(value for value in (row.get("method"), row.get("note")) if value),
            float(row.get("amount") or 0),
        )
        for col, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=col, value=value)
            cell.border = border
            cell.font = Font(name="Arial", size=9, color="1F2937")
            cell.alignment = Alignment(horizontal="right" if col == 6 else "left", vertical="top", wrap_text=True)
            if col == 6:
                cell.number_format = '#,##0.00 "EUR"'
        sheet.row_dimensions[row_index].height = 24
    if not rows:
        sheet.cell(row=2, column=1, value=report_text("empty_period", language))
    sheet.auto_filter.ref = f"A1:F{max(2, len(rows) + 1)}"
    sheet.print_title_rows = "1:1"


def _accountant_corrections_sheet(workbook: Workbook, report: dict[str, Any]) -> None:
    language = _report_language(report)
    sheet = workbook.create_sheet(report_text("sheet_corrections", language))
    headers = (report_text("date", language), report_text("type", language), report_text("document_number", language), report_text("customer", language), report_text("description", language), report_text("net_amount", language), report_text("vat", language), report_text("total", language))
    border = _accountant_sheet_style(sheet, headers, (13, 23, 24, 28, 48, 18, 16, 18))
    rows: list[dict[str, Any]] = []
    for item in report.get("credit_note_rows") or []:
        rows.append({**item, "type": report_text("credit_note", language)})
    for item in report.get("cancelled_rows") or []:
        rows.append({**item, "type": report_text("cancelled_invoice", language)})
    rows.sort(key=lambda item: (str(item.get("document_date") or ""), str(item.get("document_no") or "")))
    for row_index, row in enumerate(rows, start=2):
        values = (
            format_date(row.get("document_date")), row.get("type") or "", row.get("document_no") or "",
            row.get("partner_name") or "", row.get("description") or "", float(row.get("net_amount") or 0),
            float(row.get("vat_amount") or 0), float(row.get("gross_amount") or 0),
        )
        for col, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=col, value=value)
            cell.border = border
            cell.font = Font(name="Arial", size=9, color="1F2937")
            cell.alignment = Alignment(horizontal="right" if col >= 6 else "left", vertical="top", wrap_text=True)
            if col >= 6:
                cell.number_format = '#,##0.00 "EUR"'
        sheet.row_dimensions[row_index].height = 24
    if not rows:
        sheet.cell(row=2, column=1, value=report_text("empty_period", language))
    sheet.auto_filter.ref = f"A1:H{max(2, len(rows) + 1)}"
    sheet.print_title_rows = "1:1"


def export_project_accountant_xlsx(report: dict[str, Any], output_path: Path) -> Path:
    """Create one workbook a bookkeeper can review without navigating app screens."""
    output_path = Path(output_path)
    language = _report_language(report)
    export_project_vat_evidence_xlsx(report, output_path)
    workbook = load_workbook(output_path)
    workbook[report_text("sheet_summary", language)].title = report_text("sheet_vat", language)
    workbook[report_text("sheet_output", language)].title = report_text("sheet_outgoing", language)
    workbook[report_text("sheet_input", language)].title = report_text("sheet_incoming", language)
    _accountant_summary_sheet(workbook, report)
    _accountant_payment_sheet(workbook, report)
    _accountant_corrections_sheet(workbook, report)
    workbook.save(output_path)
    return output_path


def _accountant_pdf_table(title: str, headers: tuple[str, ...], rows: list[tuple[Any, ...]], widths: list[float], *, style_token: str, empty_message: str) -> list[Any]:
    body = _pstyle(f"accountant-body-{style_token}", 7, color=TEXT_DARK, leading=8.5)
    money = _pstyle(f"accountant-money-{style_token}", 7, color=TEXT_DARK, align=TA_RIGHT, leading=8.5)
    header = _pstyle(f"accountant-header-{style_token}", 6.8, bold=True, color=colors.white, align=TA_CENTER, leading=8)
    data: list[list[Any]] = [[_vat_pdf_cell(label, header) for label in headers]]
    for row in rows:
        data.append([_vat_pdf_cell(value, money if index == len(row) - 1 else body) for index, value in enumerate(row)])
    if not rows:
        data.append([_vat_pdf_cell(empty_message, body), *[_vat_pdf_cell("", body) for _ in headers[1:]]])
    table = Table(data, colWidths=widths, repeatRows=1, hAlign="LEFT")
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), GREEN_MID),
        ("GRID", (0, 0), (-1, -1), 0.35, GREEN_LINE),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, GREEN_LIGHTER]),
    ]))
    return [Paragraph(title, _pstyle(f"accountant-heading-{style_token}", 12, bold=True, color=GREEN_DARK)), Spacer(1, 4 * mm), table, Spacer(1, 7 * mm)]


def export_project_accountant_pdf(report: dict[str, Any], output_path: Path) -> Path:
    """Render the whole accountant package as selectable-text PDF, not as a screenshot."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    language = _report_language(report)
    project = report.get("project") if isinstance(report.get("project"), dict) else {}
    company = report.get("company") if isinstance(report.get("company"), dict) else {}
    page_size = landscape(A4)
    document = SimpleDocTemplate(
        str(output_path), pagesize=page_size, leftMargin=13 * mm, rightMargin=13 * mm,
        topMargin=12 * mm, bottomMargin=13 * mm, title=f"{report_text('accountant_title', language)} - {project.get('name') or ''}",
    )
    story: list[Any] = [
        Paragraph(report_text("accountant_title", language), _pstyle("accountant-title", 18, bold=True, color=GREEN_DARK)),
        Spacer(1, 3 * mm),
        Paragraph(
            escape(
                f"{report_text('company', language)}: {company.get('name') or '-'} | {report_text('project', language)}: {project.get('name') or '-'} | "
                f"{report_text('site', language)}: {project.get('site_address') or '-'} | "
                f"{report_text('period', language)}: {format_date(report.get('period_from'))} - {format_date(report.get('period_to'))}"
            ),
            _pstyle("accountant-detail", 8.5, color=TEXT_MUTED, leading=11),
        ),
        Spacer(1, 5 * mm),
    ]
    totals = report.get("totals") if isinstance(report.get("totals"), dict) else {}
    summary = Table(
        [
            [report_text("outgoing_invoices", language), format_currency(totals.get("output_gross") or 0, "EUR")],
            [report_text("incoming_bills", language), format_currency(totals.get("input_gross") or 0, "EUR")],
            [report_text("collected", language), format_currency(totals.get("net_collected") or 0, "EUR")],
            [report_text("vat_payable", language), format_currency(totals.get("vat_payable") or 0, "EUR")],
        ],
        colWidths=[135 * mm, 136 * mm], hAlign="LEFT",
    )
    summary.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), GREEN_LIGHT),
        ("GRID", (0, 0), (-1, -1), 0.4, GREEN_LINE),
        ("FONTNAME", (0, 0), (-1, -1), FONT_REGULAR),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.extend([summary, Spacer(1, 8 * mm)])
    story.extend(_vat_pdf_ledger(list(report.get("output_rows") or []), report_text("outgoing_invoices", language), language, "accountant-output"))
    story.extend(_vat_pdf_ledger(list(report.get("input_rows") or []), report_text("incoming_bills", language), language, "accountant-input"))
    payment_rows = [
        (
            format_date(row.get("payment_date")), _localized_document_type(row.get("type"), language), row.get("invoice_number") or "",
            row.get("partner_name") or "", " | ".join(value for value in (row.get("method"), row.get("note")) if value),
            format_currency(row.get("amount") or 0, "EUR"),
        )
        for row in report.get("payment_rows") or []
    ]
    story.extend(_accountant_pdf_table(report_text("payments_refunds", language), _accountant_payment_headers(language), payment_rows, [24 * mm, 30 * mm, 37 * mm, 48 * mm, 96 * mm, 36 * mm], style_token="accountant-payments", empty_message=report_text("empty_period", language)))
    correction_rows = [
        (
            format_date(row.get("document_date")), report_text("credit_note", language), row.get("document_no") or "",
            row.get("partner_name") or "", row.get("description") or "", format_currency(row.get("gross_amount") or 0, "EUR"),
        )
        for row in report.get("credit_note_rows") or []
    ] + [
        (
            format_date(row.get("document_date")), report_text("cancelled_invoice", language), row.get("document_no") or "",
            row.get("partner_name") or "", row.get("description") or "", format_currency(row.get("gross_amount") or 0, "EUR"),
        )
        for row in report.get("cancelled_rows") or []
    ]
    correction_section = _accountant_pdf_table(
        report_text("credit_cancellations", language), (report_text("date", language), report_text("type", language), report_text("number", language), report_text("customer", language), report_text("description", language), report_text("total", language)), correction_rows,
        [24 * mm, 34 * mm, 38 * mm, 49 * mm, 90 * mm, 36 * mm], style_token="accountant-corrections", empty_message=report_text("empty_period", language),
    )
    # Do not let the final decorative spacer create an otherwise empty last page.
    story.extend(correction_section[:-1])
    foreign = list(report.get("foreign_currency_rows") or []) + list(report.get("foreign_currency_payments") or [])
    missing = list(report.get("missing_date_rows") or [])
    if foreign or missing:
        warning = report_text("accountant_control_warning", language).format(
            foreign=len(foreign), missing=len(missing), control=report_text("sheet_control", language)
        )
        story.append(Paragraph(escape(warning), _pstyle("accountant-warning", 8.5, bold=True, color=colors.HexColor("#9A3412"))))

    def footer(c: canvas.Canvas, doc: Any) -> None:
        c.saveState()
        c.setStrokeColor(GREEN_LINE)
        c.line(13 * mm, 8 * mm, page_size[0] - 13 * mm, 8 * mm)
        c.setFont(FONT_REGULAR, 6.8)
        c.setFillColor(TEXT_MUTED)
        c.drawString(13 * mm, 4.8 * mm, report_text("accountant_footer", language))
        c.drawRightString(page_size[0] - 13 * mm, 4.8 * mm, f"{report_text('page', language)} {doc.page}")
        c.restoreState()

    document.build(story, onFirstPage=footer, onLaterPages=footer)
    return output_path


def export_project_accountant_bundle(report: dict[str, Any], output_dir: Path) -> dict[str, Path]:
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    base = _accountant_report_base_name(report)
    xlsx_path = export_project_accountant_xlsx(report, output_dir / f"{base}.xlsx")
    pdf_path = export_project_accountant_pdf(report, output_dir / f"{base}.pdf")
    return {"pdf": pdf_path, "xlsx": xlsx_path}
